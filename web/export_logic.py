# web/export_logic.py
import os
import base64
import gzip
import json
import rjsmin
import rcssmin
from flask import render_template
from markupsafe import Markup
from shared import config

def get_default_filter_values(hide_offtopic_param, year_from_param, year_to_param, min_page_count_param):
    """Extracts and validates filter parameters, returning default values if invalid/missing."""
    hide_offtopic = True
    if hide_offtopic_param is not None:
        hide_offtopic = hide_offtopic_param.lower() in ['1', 'true', 'yes', 'on']
    try: year_from_value = int(year_from_param) if year_from_param is not None else config.DEFAULT_YEAR_FROM
    except ValueError: year_from_value = config.DEFAULT_YEAR_FROM
    try: year_to_value = int(year_to_param) if year_to_param is not None else config.DEFAULT_YEAR_TO
    except ValueError: year_to_value = config.DEFAULT_YEAR_TO
    try: min_page_count_value = int(min_page_count_param) if min_page_count_param is not None else config.DEFAULT_MIN_PAGE_COUNT
    except ValueError: min_page_count_value = config.DEFAULT_MIN_PAGE_COUNT
    return hide_offtopic, year_from_value, year_to_value, min_page_count_value


def font_to_data_uri(font_path):
    """Converts a font file to a Base64 data URI string."""
    with open(font_path, 'rb') as font_file:
        font_data = font_file.read()
        base64_data = base64.b64encode(font_data).decode('ascii')
        
    # Determine format from file extension
    if font_path.lower().endswith('.woff2'):
        mime_type = 'font/woff2'
        format_str = 'woff2'
    elif font_path.lower().endswith('.woff'):
        mime_type = 'font/woff'
        format_str = 'woff'
    elif font_path.lower().endswith('.ttf'):
        mime_type = 'font/ttf'
        format_str = 'truetype'
    else:
        mime_type = 'application/octet-stream'
        format_str = 'truetype'  # fallback
        
    return f"data:{mime_type};base64,{base64_data}", format_str

def embed_fonts_in_css(static_dir):
    fonts_dir = os.path.join(static_dir, 'fonts')
    font_files = [
        ('Twemoji.mozilla.ttf', 'Twemoji Mozilla', 400, 'normal'),
        ('inter-tight-v7-latin_latin-ext-300.woff2', 'Inter Tight', 300, 'normal'),
        ('inter-tight-v7-latin_latin-ext-regular.woff2', 'Inter Tight', 400, 'normal'),
        ('inter-tight-v7-latin_latin-ext-600.woff2', 'Inter Tight', 600, 'normal'),
    ]
    css_content = "/* Embedded Fonts */\n"
    font_faces = {}
    for filename, font_family, font_weight, font_style in font_files:
        font_path = os.path.join(fonts_dir, filename)
        try:
            data_uri, format_str = font_to_data_uri(font_path)
            key = (font_family, font_weight, font_style)
            if key not in font_faces: font_faces[key] = []
            font_faces[key].append((data_uri, format_str))
        except FileNotFoundError: continue

    for (font_family, font_weight, font_style), sources in font_faces.items():
        css_content += f"\n@font-face {{\n    font-display: swap;\n    font-family: '{font_family}';\n    font-style: {font_style};\n    font-weight: {font_weight};\n    src: "
        src_parts = [f"url('{data_uri}') format('{format_str}')" for data_uri, format_str in sources]
        css_content += ",\n    ".join(src_parts) + ";\n}\n"
    return css_content


def _flatten_dict(d, parent_key=''):
    items = {}
    if not isinstance(d, dict): return {parent_key: d} if parent_key else {}
    for k, v in d.items():
        if k == 'certainty_map': continue
        new_key = f"{parent_key}.{k}" if parent_key else k
        if isinstance(v, dict): items.update(_flatten_dict(v, new_key))
        else: items[new_key] = v
    return items

def prepare_history_log_data(paper_dict, set_num=None):
    """
    Prepares llm_log data for history row rendering.
    Marks malformed entries as invalid instead of scrapping them.
    """
    if set_num and set_num in [1, 2, 3]:
        raw_log = paper_dict.get(f'set_{set_num}_llm_log', '[]')
    else:
        raw_log = paper_dict.get('llm_log', '[]')
        
    try:
        log_entries = json.loads(raw_log) if raw_log else []
    except (json.JSONDecodeError, TypeError):
        log_entries = []
        
    for entry in log_entries:
        try:
            output_raw = entry.get('output', '{}')
            if isinstance(output_raw, str) and output_raw:
                entry['output'] = json.loads(output_raw)
            elif isinstance(output_raw, dict):
                entry['output'] = output_raw
            else:
                entry['output'] = {}
        except (json.JSONDecodeError, TypeError, AttributeError):
            entry['output'] = {}
            
        # 2. Trust the Backend completely.
        # The backend is the single source of truth. It stamped 'valid' 
        # and 'invalid_reason' at the exact moment the entry was created.
        # no, there's no such a thing as "legacy data" to care about.
        entry['valid'] = bool(entry.get('valid', False))

    # PASS 1: Ascending order - Mark changed cells (valid entries only)
    # averaged_llm IS included here so changes between averaged states are highlighted.
    table_entries = [e for e in log_entries
                     if e.get('type') in ['classifier', 'consensus', 'averaged_llm', 'user']
                     and e.get('valid', False)]
                     
    for i in range(len(table_entries)):
        current = table_entries[i]
        older = table_entries[i - 1] if i > 0 else None
        current['changed_fields'] = set()
        if older:
            older_flat = _flatten_dict(older.get('output', {}) or {})
            current_flat = _flatten_dict(current.get('output', {}) or {})
            
            all_keys = set(older_flat.keys()) | set(current_flat.keys())
            for key in all_keys:
                if older_flat.get(key) != current_flat.get(key):
                    current['changed_fields'].add(key)

    changed_fields_map = {
        entry['timestamp']: entry.get('changed_fields', set())
        for entry in table_entries
    }

    # PASS 2: Reverse for UI, attach verifiers
    log_entries.reverse()
    processed_entries = []
    cached_verifier = None
    for entry in log_entries:
        entry_type = entry.get('type', '')
        entry['changed_fields'] = changed_fields_map.get(entry['timestamp'], set())
        
        if entry_type == 'verifier':
            verifier_output = entry.get('output', {}) or {}
            if not isinstance(verifier_output, dict):
                verifier_output = {}
            cached_verifier = {
                'verified': verifier_output.get('verified'),
                'estimated_score': verifier_output.get('estimated_score'),
                'verifier_trace': entry.get('trace', ''),
                'verifier_model': entry.get('model', ''),
                'verifier_timestamp': entry['timestamp']
            }
            entry['verification_data'] = None
        elif entry_type in ['classifier', 'consensus']:
            # CRITICAL RESTORATION: averaged_llm is intentionally EXCLUDED here.
            # It derives its own verified/score mathematically from the sets in recalculate_main_set.
            # It does NOT consume the verifier meant for the underlying classifier runs.
            if entry.get('valid', False):
                entry['verification_data'] = cached_verifier
                cached_verifier = None # Consume it
            else:
                entry['verification_data'] = None
        else:
            entry['verification_data'] = None
            
        processed_entries.append(entry)
        
    return processed_entries


def generate_html_export_content(papers, hide_offtopic, year_from_value, year_to_value,
                                 min_page_count_value, is_lite_export=False, skip_abstracts=False):
    for paper in papers:
        paper['llm_log_entries'] = prepare_history_log_data(paper, set_num=None)
        paper['set_1_llm_log_entries'] = prepare_history_log_data(paper, set_num=1)
        paper['set_2_llm_log_entries'] = prepare_history_log_data(paper, set_num=2)
        paper['set_3_llm_log_entries'] = prepare_history_log_data(paper, set_num=3)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    static_dir = os.path.join(script_dir, 'static')
    fonts_parent_dir = os.path.join(script_dir, 'static/css')

    domain_config = config.load_domain_config()
    
    def read_static(rel_path):
        path = os.path.join(static_dir, rel_path)
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f: return f.read()
        print(f"Warning: Static file not found: {rel_path}")
        return ""

    fonts_css_content = embed_fonts_in_css(fonts_parent_dir)
    style_css_content = read_static('css/style.css')

    # Append the theme CSS directly to the bundled stylesheet
    style_css_content = fonts_css_content + "\n" + style_css_content + "\n" + domain_config.get('theme_css', '')
    style_css_content = rcssmin.cssmin(style_css_content)

    chart_js_content = read_static('libs/chart.min.js')
    chart_js_datalabels_content = read_static('libs/chartjs-plugin-datalabels.min.js')
    d3_js_content = read_static('libs/d3.min.js')
    d3_cloud_js_content = read_static('libs/d3-cloud.min.js')
    pako_js_content = read_static('libs/pako.min.js')
    
    stats_core_js = read_static('js/stats/stats_core.js')
    stats_generic_js = read_static('js/stats/stats_generic.js')
    stats_charts_js = read_static('js/stats/stats_charts.js')
    stats_domain_js = read_static('js/stats/stats_domain.js')
    stats_latex_js = read_static('js/stats/stats_latex.js')
    
    # AFTER
    filtering_js = (
        read_static('js/filtering/filtering_state.js') + '\n' +
        read_static('js/filtering/filtering_engine.js') + '\n' +
        read_static('js/filtering/filtering_actions.js') + '\n' +
        read_static('js/filtering/filtering_init.js')
    )
    ghpages_js = read_static('js/ghpages.js') or read_static('ghpages.js')

    # Minify JS
    for var in ['chart_js_content', 'chart_js_datalabels_content', 'd3_js_content', 'd3_cloud_js_content', 
                'stats_core_js', 'stats_generic_js', 'stats_charts_js', 'stats_domain_js', 'stats_latex_js',
                'filtering_js', 'ghpages_js']:
        locals()[var] = rjsmin.jsmin(locals()[var])

    papers_table_static_export = render_template(
        'static_export/papers_table_static_export.html', papers=papers, domain_config=domain_config,
        type_emojis=config.TYPE_EMOJIS, pdf_emojis=config.PDF_EMOJIS, default_type_emoji=config.DEFAULT_TYPE_EMOJI,
        hide_offtopic=hide_offtopic, year_from_value=str(year_from_value), year_to_value=str(year_to_value),
        min_page_count_value=str(min_page_count_value), is_lite_export=is_lite_export,
        skip_abstracts=skip_abstracts, 
    )
    
    full_html_content = render_template(
        'static_export/index_static_export.html', domain_config=domain_config,
        papers_table_static_export=papers_table_static_export, hide_offtopic=hide_offtopic,
        year_from_value=year_from_value, year_to_value=year_to_value, min_page_count_value=min_page_count_value,
        # total_paper_count=len(papers),
        style_css_content=Markup(style_css_content), chart_js_content=Markup(chart_js_content),
        chart_js_datalabels_content=Markup(chart_js_datalabels_content), d3_js_content=Markup(d3_js_content),
        d3_cloud_js_content=Markup(d3_cloud_js_content), stats_core_js=Markup(stats_core_js),
        stats_generic_js=Markup(stats_generic_js), stats_charts_js=Markup(stats_charts_js),
        stats_domain_js=Markup(stats_domain_js), stats_latex_js=Markup(stats_latex_js),
        filtering_js=Markup(filtering_js), ghpages_js=Markup(ghpages_js)
    )
    
    # --- Compress the full HTML content ---
    html_bytes = full_html_content.encode('utf-8')
    compressed_bytes = gzip.compress(html_bytes)
    compressed_base64 = base64.b64encode(compressed_bytes).decode('ascii')
    
    # --- Render the LOADER template, passing the compressed data ---
    loader_html_content = render_template(
        'static_export/loader.html', compressed_html_data=compressed_base64, pako_js_content=Markup(pako_js_content)
    )
    return loader_html_content

def generate_xlsx_export_content(papers):
    """Generates a comprehensive, multi-sheet Excel workbook for domain-agnostic data export."""
    import io
    import json
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from shared import config

    wb = Workbook()
    domain_config = config.load_domain_config()
    
    # --- Helper Functions ---
    def get_val(d, path):
        """Safely traverse nested dicts using dot-notation."""
        if not d or not path: return None
        keys = path.split('.')
        for k in keys:
            if isinstance(d, dict) and k in d:
                d = d[k]
            else:
                return None
        return d

    def format_excel_value(val, key=None):
        """Converts DB/JSON values to Excel-friendly types."""
        # Force numeric for specific scoring/metadata fields
        if key in ('relevance', 'estimated_score', 'page_count', 'year', 'user_override_count'):
            if val is None or val == "": return ""
            try: return float(val) if '.' in str(val) else int(val)
            except: return val
            
        # Handle booleans
        if val is True or (isinstance(val, (int, float)) and val == 1): return True
        if val is False or (isinstance(val, (int, float)) and val == 0): return False
        if val is None or val == "": return ""
        if isinstance(val, str):
            if val.lower() in ('true', '1', 'yes'): return True
            if val.lower() in ('false', '0', 'no'): return False
        return val

    # --- Shared Styles ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(wrap_text=True, vertical='bottom', horizontal='center')
    thin_border = Border(bottom=Side(style='thin', color="FFFFFF"))
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # For conflicts
    
    table_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    # ==========================================
    # SHEET 1: PAPERS (Main Table)
    # ==========================================
    ws_papers = wb.active
    ws_papers.title = "Papers"
    
    # Dynamically build columns based on domain_config
    columns = []
    columns.append(('Metadata', 'ID', 'id', None, None))
    columns.append(('Metadata', 'Type', 'type', None, None))
    columns.append(('Metadata', 'Title', 'title', None, None))
    columns.append(('Metadata', 'Authors', 'authors', None, None))
    columns.append(('Metadata', 'Year', 'year', None, None))
    columns.append(('Metadata', 'Journal/Conference', 'journal', None, None))
    columns.append(('Metadata', 'Pages', 'page_count', None, None))
    columns.append(('Metadata', 'DOI', 'doi', None, None))
    columns.append(('Metadata', 'Keywords', 'keywords', None, None))
    
    columns.append(('Universal Inferred', 'Off-topic', 'is_offtopic', 'classification', 'main_certainty'))
    columns.append(('Universal Inferred', 'Relevance', 'relevance', 'classification', 'main_certainty'))
    
    for group in domain_config.get('groups', []):
        group_name = group.get('friendly_name', group.get('label', 'Domain Group'))
        if group.get('filter_type') == 'tri_state':
            columns.append((group_name, group.get('label', group['json_path']), group['json_path'], 'classification', 'main_certainty'))
        else:
            for field in group.get('fields', []):
                col_name = field.get('label', field['key'])
                json_path = f"{group['json_path']}.{field['key']}"
                columns.append((group_name, col_name, json_path, 'classification', 'main_certainty'))
                
    for field in domain_config.get('editable_fields', []):
        columns.append(('Editable Fields', field.get('label', field['json_path']), field['json_path'], 'classification', None))
        
    columns.append(('Verification', 'Verified', 'verified', None, 'main_certainty'))
    columns.append(('Verification', 'Est. Score', 'estimated_score', None, None))
    columns.append(('Verification', 'Verified By', 'verified_by', None, None))
    columns.append(('Audit', 'User Overrides', 'user_override_count', None, None))
    columns.append(('Audit', 'Changed By', 'changed_by', None, None))
    columns.append(('Audit', 'Last Changed', 'changed', None, None))
    columns.append(('Audit', 'User Comments', 'user_trace', None, None))
    columns.append(('Files', 'PDF State', 'pdf_state', None, None))
    columns.append(('Files', 'PDF Filename', 'pdf_filename', None, None))

    # Write Headers
    for col_idx, (group_name, col_name, *_) in enumerate(columns, 1):
        cell = ws_papers.cell(row=1, column=col_idx, value=f"{group_name}\n{col_name}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    bool_cols = []
    score_cols = []
    
    # Write Data Rows
    for row_idx, paper in enumerate(papers, 2):
        classification = paper.get('classification')
        if isinstance(classification, str):
            try: classification = json.loads(classification)
            except: classification = {}
        if not classification: classification = {}
            
        certainty = paper.get('main_certainty')
        if isinstance(certainty, str):
            try: certainty = json.loads(certainty)
            except: certainty = {}
        if not certainty: certainty = {}
        
        for col_idx, col_def in enumerate(columns, 1):
            group_name, col_name, key = col_def[0], col_def[1], col_def[2]
            source_dict = col_def[3] if len(col_def) > 3 else None
            cert_dict = col_def[4] if len(col_def) > 4 else None
            
            val = None
            if source_dict == 'classification':
                val = get_val(classification, key)
            elif source_dict is None:
                val = paper.get(key)
                
            cert = None
            if cert_dict == 'main_certainty':
                cert = get_val(certainty, key)
                
            excel_val = format_excel_value(val, key)
            
            # Format ISO timestamps to human-readable dates
            if key == 'changed' and excel_val:
                try:
                    dt = datetime.fromisoformat(str(excel_val).replace('Z', '+00:00'))
                    excel_val = dt.strftime("%Y-%m-%d %H:%M:%S")
                except:
                    pass
                    
            cell = ws_papers.cell(row=row_idx, column=col_idx, value=excel_val)
            
            # Apply Conditional Formatting based on data type and certainty
            if key not in ('relevance', 'estimated_score', 'page_count', 'year', 'user_override_count') and isinstance(excel_val, bool):
                if excel_val is True: cell.fill = green_fill
                elif excel_val is False: cell.fill = red_fill
                bool_cols.append(col_idx)
            elif excel_val == "":
                cell.fill = gray_fill
                
            if cert == 'conflict':
                cell.fill = orange_fill
                
            if key in ('relevance', 'estimated_score'):
                score_cols.append(col_idx)
                
            if key in ('title', 'abstract', 'user_trace'):
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Format as an Official Excel Table
    last_col_letter = get_column_letter(len(columns))
    last_row = len(papers) + 1
    tab = Table(displayName="PapersTable", ref=f"A1:{last_col_letter}{last_row}")
    tab.tableStyleInfo = table_style
    ws_papers.add_table(tab)
    
    # Color Scale for Scores (0 to 10)
    for col_idx in set(score_cols):
        col_letter = get_column_letter(col_idx)
        rule = ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',
            mid_type='num', mid_value=5, mid_color='FFEB84',
            end_type='num', end_value=10, end_color='63BE7B'
        )
        ws_papers.conditional_formatting.add(f"{col_letter}2:{col_letter}{last_row}", rule)

    ws_papers.freeze_panes = 'A2'
    
    # Auto-adjust column widths
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(ws_papers.cell(row=1, column=col_idx).value).split('\n')[-1])
        for row in range(2, min(last_row + 1, 50)):
            val = ws_papers.cell(row=row, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 50))
        ws_papers.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # ==========================================
    # SHEET 2: CLASSIFICATION HISTORY (Audit Log)
    # ==========================================
    ws_hist = wb.create_sheet("Classification History")
    hist_headers = ["Paper ID", "Paper Title", "Timestamp", "Type", "Model", "Valid", "Invalid Reason", "Trace (Reasoning)", "Output (JSON)", "Certainty Map (JSON)"]
    
    for col_idx, header in enumerate(hist_headers, 1):
        cell = ws_hist.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    hist_row_idx = 2
    for paper in papers:
        paper_id = paper.get('id')
        paper_title = paper.get('title', '')
        llm_log_raw = paper.get('llm_log', '[]')
        try:
            log_entries = json.loads(llm_log_raw) if llm_log_raw else []
        except:
            log_entries = []
            
        for entry in log_entries:
            ts = entry.get('timestamp', '')
            try:
                dt = datetime.fromisoformat(str(ts).replace('Z', '+00:00'))
                ts_formatted = dt.strftime("%Y-%m-%d %H:%M:%S")
            except:
                ts_formatted = ts
                
            entry_type = entry.get('type', '')
            model = entry.get('model', '')
            valid = entry.get('valid', False)
            invalid_reason = entry.get('invalid_reason', '')
            trace = entry.get('trace', '')
            
            output_raw = entry.get('output', '{}')
            if not isinstance(output_raw, str):
                output_str = json.dumps(output_raw)
            else:
                output_str = output_raw
                
            cert_map = entry.get('certainty_map', {})
            cert_str = json.dumps(cert_map) if cert_map else ''
            
            ws_hist.cell(row=hist_row_idx, column=1, value=paper_id)
            ws_hist.cell(row=hist_row_idx, column=2, value=paper_title)
            ws_hist.cell(row=hist_row_idx, column=3, value=ts_formatted)
            ws_hist.cell(row=hist_row_idx, column=4, value=entry_type)
            ws_hist.cell(row=hist_row_idx, column=5, value=model)
            ws_hist.cell(row=hist_row_idx, column=6, value="Yes" if valid else "No")
            ws_hist.cell(row=hist_row_idx, column=7, value=invalid_reason)
            ws_hist.cell(row=hist_row_idx, column=8, value=trace).alignment = Alignment(wrap_text=True, vertical='top')
            ws_hist.cell(row=hist_row_idx, column=9, value=output_str)
            ws_hist.cell(row=hist_row_idx, column=10, value=cert_str)
            
            hist_row_idx += 1

    if hist_row_idx > 2:
        hist_tab = Table(displayName="HistoryTable", ref=f"A1:{get_column_letter(len(hist_headers))}{hist_row_idx - 1}")
        hist_tab.tableStyleInfo = table_style
        ws_hist.add_table(hist_tab)
        
    ws_hist.freeze_panes = 'A2'
    for col_idx in range(1, len(hist_headers) + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 8: ws_hist.column_dimensions[col_letter].width = 80
        elif col_idx in (9, 10): ws_hist.column_dimensions[col_letter].width = 50
        else: ws_hist.column_dimensions[col_letter].width = 20

    # ==========================================
    # SHEET 3: LLM SETS COMPARISON (Averaging & Disagreements)
    # ==========================================
    ws_sets = wb.create_sheet("LLM Sets Comparison")
    sets_headers = ["Paper ID", "Paper Title", "Field Path", "Set 1", "Set 2", "Set 3", "Averaged (Main)", "Certainty"]
    for col_idx, header in enumerate(sets_headers, 1):
        cell = ws_sets.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        
    sets_row_idx = 2
    
    for paper in papers:
        paper_id = paper.get('id')
        paper_title = paper.get('title', '')
        
        def parse_json_blob(blob):
            if not blob: return {}
            if isinstance(blob, str):
                try: return json.loads(blob)
                except: return {}
            return blob
            
        s1 = parse_json_blob(paper.get('set_1_llm'))
        s2 = parse_json_blob(paper.get('set_2_llm'))
        s3 = parse_json_blob(paper.get('set_3_llm'))
        classification = parse_json_blob(paper.get('classification'))
        certainty = parse_json_blob(paper.get('main_certainty'))
        
        paper_paths = set()
        def collect_paths(d, prefix=''):
            if not isinstance(d, dict): return
            for k, v in d.items():
                path = f"{prefix}.{k}" if prefix else k
                if isinstance(v, dict):
                    collect_paths(v, path)
                else:
                    paper_paths.add(path)
                    
        # Dynamically collect all JSON paths from all sets to guarantee domain-agnostic completeness
        collect_paths(s1)
        collect_paths(s2)
        collect_paths(s3)
        collect_paths(classification)
        
        if not paper_paths:
            continue
            
        for path in sorted(list(paper_paths)):
            v1 = get_val(s1, path)
            v2 = get_val(s2, path)
            v3 = get_val(s3, path)
            v_avg = get_val(classification, path)
            cert = get_val(certainty, path)
            
            ws_sets.cell(row=sets_row_idx, column=1, value=paper_id)
            ws_sets.cell(row=sets_row_idx, column=2, value=paper_title)
            ws_sets.cell(row=sets_row_idx, column=3, value=path)
            ws_sets.cell(row=sets_row_idx, column=4, value=format_excel_value(v1, path.split('.')[-1]))
            ws_sets.cell(row=sets_row_idx, column=5, value=format_excel_value(v2, path.split('.')[-1]))
            ws_sets.cell(row=sets_row_idx, column=6, value=format_excel_value(v3, path.split('.')[-1]))
            ws_sets.cell(row=sets_row_idx, column=7, value=format_excel_value(v_avg, path.split('.')[-1]))
            ws_sets.cell(row=sets_row_idx, column=8, value=cert)
            
            sets_row_idx += 1

    if sets_row_idx > 2:
        sets_tab = Table(displayName="SetsTable", ref=f"A1:{get_column_letter(len(sets_headers))}{sets_row_idx - 1}")
        sets_tab.tableStyleInfo = table_style
        ws_sets.add_table(sets_tab)
        
    ws_sets.freeze_panes = 'D2'
    for col_idx in range(1, len(sets_headers) + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx in (1, 2, 3): ws_sets.column_dimensions[col_letter].width = 25
        else: ws_sets.column_dimensions[col_letter].width = 15

    # Finalize and return bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_filename(base_name, year_from, year_to, min_page_count, hide_offtopic, extra_suffix=""):
    filename_parts = [base_name]
    if year_from == year_to: filename_parts.append(str(year_from))
    else: filename_parts.append(f"{year_from}-{year_to}")
    if min_page_count > 0: filename_parts.append(f"min{min_page_count}pg")
    if hide_offtopic: filename_parts.append("noOfftopic")
    if extra_suffix: filename_parts.append(extra_suffix)
    return "_".join(filename_parts)
