import json
from io import BytesIO

import pytest
from openpyxl import load_workbook

from shared import db
from web import export_logic


@pytest.fixture()
def mock_domain_config(monkeypatch):
    def _load():
        return {
            'domain_name': 'Test',
            'groups': [
                {'name': 'g1', 'filter_type': 'tri_state', 'json_path': 'is_survey', 'label': 'Survey'},
                {'name': 'g2', 'filter_type': 'inclusion', 'json_path': 'features', 'friendly_name': 'Features',
                 'fields': [{'key': 'smt', 'label': 'SMT'}]}
            ],
            'editable_fields': [],
            'theme': {}
        }
    monkeypatch.setattr(export_logic.config, 'load_domain_config', _load)
    yield

def test_excel_dynamic_columns_and_conflict_highlight(test_db, mock_domain_config):
    with db.get_db() as conn:
        conn.execute("""
            INSERT INTO papers (id, title, classification, main_certainty) 
            VALUES ('p1', 'T', ?, ?)
        """, (
            json.dumps({"is_survey": True, "features": {"smt": False}}),
            json.dumps({"is_survey": "solid", "features.smt": "conflict"})
        ))
    
    papers = db.fetch_papers(hide_offtopic=False)
    xlsx_bytes = export_logic.generate_xlsx_export_content(papers)
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb['Papers']
    
    smt_col_idx = None
    for col in range(1, ws.max_column + 1):
        if 'SMT' in str(ws.cell(row=1, column=col).value):
            smt_col_idx = col
            break
            
    assert smt_col_idx is not None, "Dynamic inclusion column 'SMT' missing"
    
    cell = ws.cell(row=2, column=smt_col_idx)
    assert cell.value is False
    # FFEB9C is the orange_fill applied to conflicts in export_logic.py
    assert cell.fill.start_color.rgb in ('00FFEB9C', 'FFEB9C')