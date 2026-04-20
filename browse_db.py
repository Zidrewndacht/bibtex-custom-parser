# browse_db.py
import sqlite3
import json
import argparse
from datetime import datetime
from flask import Flask, render_template, request, jsonify, abort, send_from_directory, Response, send_file
from markupsafe import Markup 
import tempfile
import os
import sys
import threading
import webbrowser
import rjsmin
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill 
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.utils import secure_filename 
import gzip
import base64
import zstandard as zstd
import tarfile
import shutil
# import subprocess
import logging
import requests
from datetime import datetime, timezone
import globals

# Define default year range - For this app:
DEFAULT_YEAR_FROM = 2016
DEFAULT_YEAR_TO = 2025
DEFAULT_MIN_PAGE_COUNT = 4

app = Flask(__name__)
logging.getLogger('werkzeug').setLevel(logging.WARNING)
DATABASE = None # Will be set from command line argument

def render_papers_table(hide_offtopic_param=None, year_from_param=None, year_to_param=None, min_page_count_param=None):
    """Fetches papers based on filters and renders the papers_table.html template. 
       Used for initial render from / and XHR updates."""
    # Determine hide_offtopic state
    hide_offtopic = True # Default
    if hide_offtopic_param is not None:
        hide_offtopic = hide_offtopic_param.lower() in ['1', 'true', 'yes', 'on']

    # Determine filter values, using defaults if not provided or invalid
    year_from_value = int(year_from_param) if year_from_param is not None else DEFAULT_YEAR_FROM
    year_to_value = int(year_to_param) if year_to_param is not None else DEFAULT_YEAR_TO
    min_page_count_value = int(min_page_count_param) if min_page_count_param is not None else DEFAULT_MIN_PAGE_COUNT

    # Fetch papers with ALL the filters applied
    papers = fetch_papers(
        hide_offtopic=hide_offtopic,
        year_from=year_from_value,
        year_to=year_to_value,
        min_page_count=min_page_count_value,
    )

    # Render the table template fragment, passing the search query value for the input field
    rendered_table = render_template(
        'papers_table.html',
        papers=papers,
        type_emojis=globals.TYPE_EMOJIS,
        default_type_emoji=globals.DEFAULT_TYPE_EMOJI,
        pdf_emojis=globals.PDF_EMOJIS, # Pass the PDF emojis dictionary
        hide_offtopic=hide_offtopic,
        # Pass the *string representations* of the values to the template for input fields
        year_from_value=str(year_from_value),
        year_to_value=str(year_to_value),
        min_page_count_value=str(min_page_count_value)
    )
    return rendered_table

# DB functions - should not be moved away from Flask process here:
def get_db_connection():
    """Create a connection to the SQLite database and ensure FTS tables."""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row 
    return conn


def fetch_papers(hide_offtopic=True, year_from=None, year_to=None, min_page_count=None):
    """Fetch papers from the database, applying various optional filters."""
    conn = get_db_connection()
    base_query = "SELECT p.* FROM papers p"
    conditions = []
    params = []
    
    if hide_offtopic:
        conditions.append("(p.is_offtopic = 0 OR p.is_offtopic IS NULL)")
    
    if year_from is not None:
        try:
            year_from = int(year_from)
            conditions.append("p.year >= ?")
            params.append(year_from)
        except (ValueError, TypeError):
            pass
    
    if year_to is not None:
        try:
            year_to = int(year_to)
            conditions.append("p.year <= ?")
            params.append(year_to)
        except (ValueError, TypeError):
            pass
    
    if min_page_count is not None:
        try:
            min_page_count = int(min_page_count)
            conditions.append("(p.page_count IS NULL OR p.page_count = '' OR p.page_count >= ?)")
            params.append(min_page_count)
        except (ValueError, TypeError):
            pass
    
    # Build Final Query
    query_parts = [base_query]
    if conditions:
        query_parts.append("WHERE " + " AND ".join(conditions))
    query_parts.append("ORDER BY (p.user_trace IS NULL OR p.user_trace = '') ASC")
    
    query = " ".join(query_parts)
    
    try:
        papers = conn.execute(query, params).fetchall()
    except sqlite3.Error as e:
        conn.close()
        print(f"Database error during fetch_papers: {e}")
        raise
    finally:
        conn.close()
    
    # Process Results
    paper_list = []
    for paper in papers:
        paper_dict = dict(paper)
        
        # Parse main JSON fields
        try:
            paper_dict['features'] = json.loads(paper_dict['features']) if paper_dict['features'] else {}
        except (json.JSONDecodeError, TypeError):
            paper_dict['features'] = {}
        
        try:
            paper_dict['technique'] = json.loads(paper_dict['technique']) if paper_dict['technique'] else {}
        except (json.JSONDecodeError, TypeError):
            paper_dict['technique'] = {}
        
        # Parse main_certainty
        try:
            paper_dict['main_certainty'] = json.loads(paper_dict['main_certainty']) if paper_dict['main_certainty'] else {}
        except (json.JSONDecodeError, TypeError):
            paper_dict['main_certainty'] = {}
        
        paper_dict['pdf_filename'] = paper_dict.get('pdf_filename')
        paper_dict['pdf_state'] = paper_dict.get('pdf_state', 'none')
        paper_dict['changed_formatted'] = format_changed_timestamp(paper_dict.get('changed'))
        
        paper_list.append(paper_dict)
    
    return paper_list
# In browse_db.py, update update_paper_custom_fields() function

def update_paper_custom_fields(paper_id, data, changed_by="user"):
    """
    Update the custom classification fields for a paper.
    User changes only affect main columns, not set_* columns.
    Updates certainty_map for all changed fields to 'solid'.
    Handles verified_by field updates.
    Verification Reset Rules:
    - If user explicitly sets verified/verified_by, those values stick
    - If user changes non-verification fields on LLM-verified paper, verification resets
    - If user changes non-verification fields on user-verified paper, verification stays
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    changed_timestamp = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
    
    # Fetch current state
    cursor.execute("""
    SELECT llm_log, user_override_count,
    last_llm_features, last_llm_technique,
    last_llm_is_survey, last_llm_is_offtopic,
    last_llm_is_through_hole, last_llm_is_smt, last_llm_is_x_ray,
    last_llm_relevance, last_llm_verified, last_llm_estimated_score,
    features, technique,
    is_survey, is_offtopic, is_through_hole, is_smt, is_x_ray,
    relevance, verified, estimated_score,
    user_trace, main_certainty, verified_by, pdf_filename, pdf_state
    FROM papers WHERE id = ?
    """, (paper_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return {'status': 'error', 'message': 'Paper not found'}
    
    (current_llm_log_str, current_user_override_count,
    last_llm_features_str, last_llm_technique_str,
    last_llm_is_survey, last_llm_is_offtopic,
    last_llm_is_through_hole, last_llm_is_smt, last_llm_is_x_ray,
    last_llm_relevance, last_llm_verified, last_llm_estimated_score,
    current_features_str, current_technique_str,
    current_is_survey, current_is_offtopic,
    current_is_through_hole, current_is_smt, current_is_x_ray,
    current_relevance, current_verified, current_estimated_score,
    current_user_trace, current_certainty_str, current_verified_by,
    current_pdf_filename, current_pdf_state) = row
    
    # === Save original verified_by BEFORE any processing ===
    original_verified_by = current_verified_by
    
    # Parse existing log
    try:
        existing_log = json.loads(current_llm_log_str) if current_llm_log_str else []
    except json.JSONDecodeError:
        existing_log = []
    
    # Parse current features/technique
    try:
        current_features = json.loads(current_features_str) if current_features_str else {}
    except:
        current_features = {}
    try:
        current_technique = json.loads(current_technique_str) if current_technique_str else {}
    except:
        current_technique = {}
    
    # Parse current certainty map
    try:
        certainty_map = json.loads(current_certainty_str) if current_certainty_str else {}
    except:
        certainty_map = {}
    
    # Prepare Database Updates
    update_fields = []
    update_values = []
    
    # Track which fields are being updated
    updated_main_fields = {}
    
    # Track if user explicitly set verification fields (to prevent reset logic)
    user_set_verified = 'verified' in data
    user_set_verified_by = 'verified_by' in data
    
    # Handle features_ updates
    feature_updates = {}
    for key in list(data.keys()):
        if key.startswith('features_'):
            feature_key = key.split('features_', 1)[1]
            value = data[key]
            if isinstance(value, str):
                if value == 'true':
                    feature_updates[feature_key] = True
                elif value == 'false':
                    feature_updates[feature_key] = False
                else:
                    feature_updates[feature_key] = None
            else:
                feature_updates[feature_key] = value
            data.pop(key)
    
    if feature_updates:
        current_features.update(feature_updates)
        update_fields.append("features = ?")
        update_values.append(json.dumps(current_features))
        for feature_key in feature_updates.keys():
            certainty_map[f'features_{feature_key}'] = 'solid'
    
    # Handle technique_ updates
    technique_updates = {}
    for key in list(data.keys()):
        if key.startswith('technique_'):
            tech_key = key.split('technique_', 1)[1]
            value = data[key]
            if isinstance(value, str):
                if value == 'true':
                    technique_updates[tech_key] = True
                elif value == 'false':
                    technique_updates[tech_key] = False
                else:
                    technique_updates[tech_key] = None
            else:
                technique_updates[tech_key] = value
            data.pop(key)
    
    if technique_updates:
        current_technique.update(technique_updates)
        update_fields.append("technique = ?")
        update_values.append(json.dumps(current_technique))
        for tech_key in technique_updates.keys():
            certainty_map[f'technique_{tech_key}'] = 'solid'
    
    # Handle main boolean fields
    main_bool_fields = ['is_survey', 'is_offtopic', 'is_through_hole', 'is_smt', 'is_x_ray']
    for field in main_bool_fields:
        if field in data:
            value = data[field]
            if isinstance(value, str):
                db_val = 1 if value.lower() in ('true', '1', 'on') else (0 if value.lower() in ('false', '0') else None)
            else:
                db_val = 1 if value is True else (0 if value is False else None)
            update_fields.append(f"{field} = ?")
            update_values.append(db_val)
            updated_main_fields[field] = db_val
            certainty_map[field] = 'solid'
            if field == 'is_survey':
                current_is_survey = db_val
            elif field == 'is_offtopic':
                current_is_offtopic = db_val
            elif field == 'is_through_hole':
                current_is_through_hole = db_val
            elif field == 'is_smt':
                current_is_smt = db_val
            elif field == 'is_x_ray':
                current_is_x_ray = db_val
            data.pop(field)
    
    # Handle other fields
    if 'research_area' in data:
        update_fields.append("research_area = ?")
        update_values.append(data['research_area'])
        data.pop('research_area')
    
    if 'page_count' in data:
        page_count_value = data['page_count']
        if page_count_value is not None:
            try:
                page_count_value = int(page_count_value)
            except (ValueError, TypeError):
                page_count_value = None
        update_fields.append("page_count = ?")
        update_values.append(page_count_value)
        data.pop('page_count')
    
    if 'relevance' in data:
        update_fields.append("relevance = ?")
        update_values.append(data['relevance'])
        current_relevance = data['relevance']
        data.pop('relevance')
    
    if 'user_trace' in data:
        update_fields.append("user_trace = ?")
        update_values.append(data['user_trace'])
        current_user_trace = data['user_trace']
        data.pop('user_trace')

        # === Automated paywall detection (Bidirectional) ===
        is_paywalled_present = 'paywalled' in str(current_user_trace).lower()
        has_pdf = bool(current_pdf_filename)

        if is_paywalled_present:
            # Keyword exists: Set to paywalled ONLY if no PDF is attached and not already set
            if not has_pdf and current_pdf_state != "paywalled":
                update_fields.append("pdf_state = ?")
                update_values.append("paywalled")
        else:
            # Keyword removed: Clear paywalled state if it was set and no PDF exists
            if not has_pdf and current_pdf_state == "paywalled":
                update_fields.append("pdf_state = ?")
                update_values.append("none")

    # === Handle verified field ===
    if 'verified' in data:
        val = data['verified']
        if isinstance(val, str):
            db_val = 1 if val.lower() in ('true', '1', 'on') else (0 if val.lower() in ('false', '0') else None)
        else:
            db_val = 1 if val is True else (0 if val is False else None)
        update_fields.append("verified = ?")
        update_values.append(db_val)
        current_verified = db_val
        # === When user explicitly sets verified, set verified_by to 'user' ===
        current_verified_by = 'user'
        update_fields.append("verified_by = ?")
        update_values.append('user')
        data.pop('verified')
    
    # === Handle verified_by field (only if user explicitly sets it separately) ===
    if 'verified_by' in data:
        verified_by_value = data['verified_by']
        db_val = None if verified_by_value == 'unknown' or verified_by_value is None else verified_by_value
        update_fields.append("verified_by = ?")
        update_values.append(db_val)
        current_verified_by = db_val
        data.pop('verified_by')
    
    if 'estimated_score' in data:
        val = data['estimated_score']
        db_val = max(0, min(100, int(val))) if isinstance(val, (int, float)) else None
        update_fields.append("estimated_score = ?")
        update_values.append(db_val)
        current_estimated_score = db_val
        data.pop('estimated_score')
    
    # === Verification Reset Logic (AFTER all field processing) ===
    # Only reset if:
    # 1. Changed by user
    # 2. User did NOT explicitly set verified or verified_by
    # 3. Paper was LLM-verified (not user-verified, not empty)
    if changed_by == "user" and not user_set_verified and not user_set_verified_by:
        was_llm_verified = (original_verified_by and
                          original_verified_by.strip() != '' and
                          original_verified_by != 'user')
        if was_llm_verified:
            # Reset verification
            update_fields.append("verified = ?")
            update_values.append(None)
            update_fields.append("estimated_score = ?")
            update_values.append(None)
            update_fields.append("verified_by = ?")
            update_values.append("")
            # === Update local variables for log entry ===
            current_verified = None
            current_estimated_score = None
            current_verified_by = ""
    
    # Calculate user_override_count
    BOOLEAN_FEATURE_KEYS = [
        'tracks', 'holes', 'bare_pcb_other', 'solder_insufficient',
        'solder_excess', 'solder_void', 'solder_crack', 'solder_other',
        'orientation', 'wrong_component', 'missing_component',
        'component_other', 'cosmetic'
    ]
    BOOLEAN_TECHNIQUE_KEYS = [
        'classic_cv_based', 'ml_traditional', 'dl_cnn_classifier',
        'dl_cnn_detector', 'dl_rcnn_detector', 'dl_transformer',
        'dl_other', 'hybrid', 'available_dataset'
    ]
    
    def normalize_bool(val):
        if val is None or val == '' or val == 'null':
            return None
        if val is True or val == 1 or val == '1' or val == 'true':
            return 1
        if val is False or val == 0 or val == '0' or val == 'false':
            return 0
        return val
    
    user_override_count = 0
    try:
        last_llm_features = json.loads(last_llm_features_str) if last_llm_features_str else {}
    except:
        last_llm_features = {}
    try:
        last_llm_technique = json.loads(last_llm_technique_str) if last_llm_technique_str else {}
    except:
        last_llm_technique = {}
    
    for key in BOOLEAN_FEATURE_KEYS:
        current_val = normalize_bool(current_features.get(key))
        llm_val = normalize_bool(last_llm_features.get(key))
        if current_val != llm_val:
            user_override_count += 1
    
    for key in BOOLEAN_TECHNIQUE_KEYS:
        current_val = normalize_bool(current_technique.get(key))
        llm_val = normalize_bool(last_llm_technique.get(key))
        if current_val != llm_val:
            user_override_count += 1
    
    for field in ['is_survey', 'is_offtopic', 'is_through_hole', 'is_smt', 'is_x_ray']:
        current_val = normalize_bool(locals()[f'current_{field}'])
        llm_val = normalize_bool(locals()[f'last_llm_{field}'])
        if current_val != llm_val:
            user_override_count += 1
    
    update_fields.append("changed = ?")
    update_values.append(changed_timestamp)
    update_fields.append("changed_by = ?")
    update_values.append(changed_by)
    update_fields.append("user_override_count = ?")
    update_values.append(user_override_count)
    update_fields.append("main_certainty = ?")
    update_values.append(json.dumps(certainty_map))
    
    # === Build user_log_output AFTER all field processing ===
    # This ensures we capture the FINAL state including verification changes
    def db_to_bool(val):
        if val == 1:
            return True
        elif val == 0:
            return False
        else:
            return None
    
    # === FULL STATE SNAPSHOT (including verification) ===
    user_log_output = {
        "is_offtopic": db_to_bool(updated_main_fields.get('is_offtopic', current_is_offtopic)),
        "relevance": current_relevance,
        "is_survey": db_to_bool(updated_main_fields.get('is_survey', current_is_survey)),
        "is_through_hole": db_to_bool(updated_main_fields.get('is_through_hole', current_is_through_hole)),
        "is_smt": db_to_bool(updated_main_fields.get('is_smt', current_is_smt)),
        "is_x_ray": db_to_bool(updated_main_fields.get('is_x_ray', current_is_x_ray)),
        "features": current_features if current_features else {},
        "technique": current_technique if current_technique else {},
        "verified": db_to_bool(current_verified),  # ← Uses UPDATED value
        "estimated_score": current_estimated_score
    }
    
    # === If verification was reset, update user_log_output to reflect it ===
    if changed_by == "user" and not user_set_verified and not user_set_verified_by:
        was_llm_verified = (original_verified_by and
                          original_verified_by.strip() != '' and
                          original_verified_by != 'user')
        if was_llm_verified:
            user_log_output['verified'] = None
            user_log_output['estimated_score'] = None
    
    user_log_entry = {
        "timestamp": changed_timestamp,
        "type": "user",
        "model": "user",
        "trace": current_user_trace or "",
        "output": json.dumps(user_log_output),
        "valid": True,
        "certainty_map": certainty_map
    }
    
    if existing_log and existing_log[-1].get('type') == 'user':
        last_user_entry = existing_log[-1]
        last_user_entry['output'] = json.dumps(user_log_output)
        last_user_entry['trace'] = current_user_trace or ""
        last_user_entry['timestamp'] = changed_timestamp
        last_user_entry['certainty_map'] = certainty_map
    else:
        existing_log.append(user_log_entry)
    
    update_fields.append("llm_log = ?")
    update_values.append(json.dumps(existing_log))
    
    if update_fields:
        update_values.append(paper_id)
        update_query = f"UPDATE papers SET {', '.join(update_fields)} WHERE id = ?"
        cursor.execute(update_query, update_values)
        conn.commit()
        rows_affected = cursor.rowcount
    else:
        rows_affected = 0
    
    conn.close()
    
    if rows_affected > 0:
        updated_paper = globals.get_paper_by_id(DATABASE, paper_id)
        if updated_paper:
            updated_paper['changed_formatted'] = format_changed_timestamp(updated_paper.get('changed'))
            return_data = {
                'status': 'success',
                'changed': updated_paper.get('changed'),
                'changed_formatted': updated_paper.get('changed_formatted'),
                'changed_by': updated_paper.get('changed_by'),
                'research_area': updated_paper.get('research_area'),
                'page_count': updated_paper.get('page_count'),
                'is_survey': updated_paper.get('is_survey'),
                'is_offtopic': updated_paper.get('is_offtopic'),
                'is_through_hole': updated_paper.get('is_through_hole'),
                'is_smt': updated_paper.get('is_smt'),
                'is_x_ray': updated_paper.get('is_x_ray'),
                'relevance': updated_paper.get('relevance'),
                'features': updated_paper.get('features', {}),
                'technique': updated_paper.get('technique', {}),
                'user_trace': updated_paper.get('user_trace'),
                'user_override_count': updated_paper.get('user_override_count'),
                'verified': updated_paper.get('verified'),
                'verified_by': updated_paper.get('verified_by'),
                'estimated_score': updated_paper.get('estimated_score'),
                'main_certainty': certainty_map,
                'pdf_state': updated_paper.get('pdf_state'),
                'pdf_filename': updated_paper.get('pdf_filename')
            }
            return return_data
        else:
            return {'status': 'error', 'message': 'Paper not found after update.'}
    else:
        return {'status': 'error', 'message': 'No rows updated.'}
    
def fetch_updated_paper_data(paper_id):
    """Fetches the full paper data after classification/verification for client-side update."""
    conn = get_db_connection()
    try:
        updated_paper = conn.execute("SELECT * FROM papers WHERE id = ?", (paper_id,)).fetchone()
        if updated_paper:
            updated_dict = dict(updated_paper)
            try:
                updated_dict['features'] = json.loads(updated_dict['features'])
            except (json.JSONDecodeError, TypeError):
                updated_dict['features'] = {}
            try:
                updated_dict['technique'] = json.loads(updated_dict['technique'])
            except (json.JSONDecodeError, TypeError):
                updated_dict['technique'] = {}
            updated_dict['changed_formatted'] = format_changed_timestamp(updated_dict.get('changed'))
            
            # Prepare data for frontend refresh
            return_data = {
                'status': 'success',
                'changed': updated_dict.get('changed'),
                'changed_formatted': updated_dict['changed_formatted'],
                'changed_by': updated_dict.get('changed_by'),
                'verified_by': updated_dict.get('verified_by'),
                'research_area': updated_dict.get('research_area'),
                'page_count': updated_dict.get('page_count'),
                'is_survey': updated_dict.get('is_survey'),
                'is_offtopic': updated_dict.get('is_offtopic'),
                'is_through_hole': updated_dict.get('is_through_hole'),
                'is_smt': updated_dict.get('is_smt'),
                'is_x_ray': updated_dict.get('is_x_ray'),
                'relevance': updated_dict.get('relevance'),
                'verified': updated_dict.get('verified'),
                'estimated_score': updated_dict.get('estimated_score'),
                'features': updated_dict['features'],
                'technique': updated_dict['technique'],
                'user_trace': updated_dict.get('user_trace'),
                'user_override_count': updated_dict.get('user_override_count')
                # REMOVED: 'reasoning_trace' and 'verifier_trace' - no longer used
            }
            return return_data
        else:
            return {'status': 'error', 'message': 'Paper not found after update.'}
    finally:
        conn.close()

def format_changed_timestamp(changed_str):
    """Format the ISO timestamp string to dd/mm/yy hh:mm:ss"""
    if not changed_str:
        return ""
    try:
        dt = datetime.fromisoformat(changed_str.replace('Z', '+00:00'))
        return dt.strftime("%d/%m/%y %H:%M:%S")
    except ValueError:
        return changed_str  # If parsing fails, return the original string


# Helpers used for HTML and XLSX exports: 
def get_default_filter_values(hide_offtopic_param, year_from_param, year_to_param, min_page_count_param):
    """Extracts and validates filter parameters, returning default values if invalid/missing."""
    hide_offtopic = True 
    if hide_offtopic_param is not None:
        hide_offtopic = hide_offtopic_param.lower() in ['1', 'true', 'yes', 'on']

    try:
        year_from_value = int(year_from_param) if year_from_param is not None else DEFAULT_YEAR_FROM
    except ValueError:
        year_from_value = DEFAULT_YEAR_FROM

    try:
        year_to_value = int(year_to_param) if year_to_param is not None else DEFAULT_YEAR_TO
    except ValueError:
        year_to_value = DEFAULT_YEAR_TO

    try:
        min_page_count_value = int(min_page_count_param) if min_page_count_param is not None else DEFAULT_MIN_PAGE_COUNT
    except ValueError:
        min_page_count_value = DEFAULT_MIN_PAGE_COUNT

    return hide_offtopic, year_from_value, year_to_value, min_page_count_value

def font_to_data_uri(font_path):
    """Converts a font file to a Base64 data URI string."""
    with open(font_path, 'rb') as font_file:
        font_data = font_file.read()
        base64_data = base64.b64encode(font_data).decode('ascii')
        
    # Determine format from file extension
    if font_path.lower().endswith('.woff2'):
        mime_type = 'font/woff2'
        format_str = 'woff2'
    elif font_path.lower().endswith('.woff'):
        mime_type = 'font/woff'
        format_str = 'woff'
    elif font_path.lower().endswith('.ttf'):
        mime_type = 'font/ttf'
        format_str = 'truetype'
    else:
        mime_type = 'application/octet-stream'
        format_str = 'truetype'  # fallback
        
    return f"data:{mime_type};base64,{base64_data}", format_str

def embed_fonts_in_css(static_dir):
    """Convert font files to Base64 data URIs and return CSS with embedded fonts."""
    fonts_dir = os.path.join(static_dir, 'fonts')
    
    # Define font files to embed
    font_files = [
        ('Twemoji.mozilla.ttf', 'Twemoji Mozilla', 400, 'normal'),
        ('inter-tight-v7-latin_latin-ext-300.woff2', 'Inter Tight', 300, 'normal'),
        ('inter-tight-v7-latin_latin-ext-regular.woff2', 'Inter Tight', 400, 'normal'),
        ('inter-tight-v7-latin_latin-ext-600.woff2', 'Inter Tight', 600, 'normal'),
    ]
    
    css_content = "/* Embedded Fonts */\n"
    
    # Group fonts by family to create proper @font-face rules
    font_faces = {}
    
    for filename, font_family, font_weight, font_style in font_files:
        font_path = os.path.join(fonts_dir, filename)
        try:
            data_uri, format_str = font_to_data_uri(font_path)
            
            key = (font_family, font_weight, font_style)
            if key not in font_faces:
                font_faces[key] = []
            
            font_faces[key].append((data_uri, format_str))
        except FileNotFoundError:
            print(f"Warning: Font file not found: {font_path}")
            continue
    
    # Generate @font-face CSS rules
    for (font_family, font_weight, font_style), sources in font_faces.items():
        css_content += f"""
/* {font_family} - {font_weight} */
@font-face {{
    font-display: swap;
    font-family: '{font_family}';
    font-style: {font_style};
    font-weight: {font_weight};
    src: """
        
        # Add all sources for this font face
        src_parts = []
        for data_uri, format_str in sources:
            src_parts.append(f"url('{data_uri}') format('{format_str}')")
        
        css_content += ",\n        ".join(src_parts) + ";\n}\n"
    
    return css_content

# Core Export Generation Functions
def generate_html_export_content(papers, hide_offtopic, year_from_value, year_to_value, min_page_count_value, is_lite_export=False):
    """Generates the full HTML content string for the static export."""
     
    # Prepare history log data for the FILTERED papers only (already passed in)
    for paper in papers:
        paper['llm_log_entries'] = prepare_history_log_data(paper, set_num=None)
        paper['set_1_llm_log_entries'] = prepare_history_log_data(paper, set_num=1)
        paper['set_2_llm_log_entries'] = prepare_history_log_data(paper, set_num=2)
        paper['set_3_llm_log_entries'] = prepare_history_log_data(paper, set_num=3)

    
    style_css_content = ""
    chart_js_content = ""
    chart_js_datalabels_content = ""
    d3_js_content = ""
    d3_cloud_js_content = ""
    ghpages_js_content = ""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    static_dir = os.path.join(script_dir, 'static')
    
    # Load and embed fonts as Base64 data URIs
    fonts_css_content = embed_fonts_in_css(static_dir)
    
    with open(os.path.join(static_dir, 'libs/chart.min.js'), 'r', encoding='utf-8') as f:
        chart_js_content = f.read()
    with open(os.path.join(static_dir, 'libs/chartjs-plugin-datalabels.min.js'), 'r', encoding='utf-8') as f:
        chart_js_datalabels_content = f.read()
    with open(os.path.join(static_dir, 'libs/d3.min.js'), 'r', encoding='utf-8') as f:
        d3_js_content = f.read()
    with open(os.path.join(static_dir, 'libs/d3-cloud.min.js'), 'r', encoding='utf-8') as f:
        d3_cloud_js_content = f.read()

    with open(os.path.join(static_dir, 'style.css'), 'r', encoding='utf-8') as f:
        style_css_content = f.read()
    with open(os.path.join(static_dir, 'ghpages.js'), 'r', encoding='utf-8') as f:
        ghpages_js_content = f.read()
    with open(os.path.join(static_dir, 'stats.js'), 'r', encoding='utf-8') as f:
        stats_js_content = f.read()
    with open(os.path.join(static_dir, 'filtering.js'), 'r', encoding='utf-8') as f:
        filtering_js_content = f.read()

    # Combine fonts CSS with main CSS
    style_css_content = fonts_css_content + "\n" + style_css_content
    
    # style_css_content = rcssmin.cssmin(style_css_content)
    
    chart_js_content = rjsmin.jsmin(chart_js_content)
    chart_js_datalabels_content = rjsmin.jsmin(chart_js_datalabels_content)
    d3_js_content = rjsmin.jsmin(d3_js_content)
    d3_cloud_js_content = rjsmin.jsmin(d3_cloud_js_content)

    stats_js_content = rjsmin.jsmin(stats_js_content)
    filtering_js_content = rjsmin.jsmin(filtering_js_content)
    ghpages_js_content = rjsmin.jsmin(ghpages_js_content)

    # --- Render the static export template ---
    papers_table_static_export = render_template(
        'papers_table_static_export.html',
        papers=papers,
        type_emojis=globals.TYPE_EMOJIS,
        pdf_emojis=globals.PDF_EMOJIS,
        default_type_emoji=globals.DEFAULT_TYPE_EMOJI,
        hide_offtopic=hide_offtopic,
        year_from_value=str(year_from_value),
        year_to_value=str(year_to_value),
        min_page_count_value=str(min_page_count_value),
        is_lite_export=is_lite_export,
    )
    full_html_content = render_template(
        'index_static_export.html',
        papers_table_static_export=papers_table_static_export,
        hide_offtopic=hide_offtopic,
        year_from_value=year_from_value,
        year_to_value=year_to_value,
        min_page_count_value=min_page_count_value,

        style_css_content=Markup(style_css_content),
        
        chart_js_content=Markup(chart_js_content),
        chart_js_datalabels_content=Markup(chart_js_datalabels_content),
        d3_js_content=Markup(d3_js_content),
        d3_cloud_js_content=Markup(d3_cloud_js_content),

        filtering_js_content=Markup(filtering_js_content),
        stats_js_content=Markup(stats_js_content),
        ghpages_js_content=Markup(ghpages_js_content)
    )

    # --- Compress the full HTML content ---
    html_bytes = full_html_content.encode('utf-8')  # 1. Encode the HTML string to bytes (UTF-8)
    compressed_bytes = gzip.compress(html_bytes)    # 2. Compress the bytes
    compressed_base64 = base64.b64encode(compressed_bytes).decode('ascii')  # 3. Encode the compressed bytes to Base64 for embedding in JS

    pako_js_content = ""
    with open(os.path.join(static_dir, 'libs/pako.min.js'), 'r', encoding='utf-8') as f:
        pako_js_content = f.read()

    # --- Render the LOADER template, passing the compressed data ---
    loader_html_content = render_template(
        'loader.html',
        compressed_html_data=compressed_base64,
        pako_js_content=Markup(pako_js_content)
    )
    return loader_html_content

def generate_xlsx_export_content(papers): #outdated, must be updated to be useful for v1.2:
    # """Generates the Excel file content as bytes."""
    # output = io.BytesIO()
    # wb = Workbook()
    # ws = wb.active
    # ws.title = "PCB Inspection Papers"

    # # --- Define Headers (Updated Order - Corrected Boolean Features) ---
    # headers = [
    #     "Type", "Title", "Year", "Journal/Conf name", "Pages count",
    #     # Classification Summary
    #     "Off-topic", "Relevance", "Survey", "THT", "SMT", "X-Ray",
    #     # Features Summary (Updated Order - Corrected Boolean Features)
    #     "Tracks", "Holes / Vias", "Bare PCB Other", # Boolean (e.g., bare_pcb_other)
    #     "Solder Insufficient", "Solder Excess", "Solder Void", "Solder Crack", "Solder Other", # Boolean (e.g., solder_other)
    #     "Missing Comp", "Wrong Comp", "Orientation", "Comp Other", # Boolean (e.g., component_other)
    #     "Cosmetic", "Other State", # Boolean for state (based on 'other' text content)
    #     "Other Defects Text", # Text for content (the 'other' field)
    #     # Techniques Summary (Updated Order)
    #     "Classic CV", "ML", "CNN Classifier", "CNN Detector",
    #     "R-CNN Detector", "Transformers", "Other DL", "Hybrid", "Datasets", "Model name",
    #     # Metadata
    #     "Last Changed", "Changed By", "Verified", "Accr. Score", "Verified By",
    #     "User Comment State", "User Comments" # Boolean for state, Text for content
    # ]

    # # --- Write Headers ---
    # for col_num, header in enumerate(headers, 1):
    #     cell = ws.cell(row=1, column=col_num, value=header)
    #     cell.font = Font(bold=True)

    # # --- Write Data Rows ---
    # for row_num, paper in enumerate(papers, 2): # Start from row 2
    #     # --- Helper function for consistent Excel value conversion ---
    #     def format_excel_value(val):
    #         """
    #         Converts Python/DB values to Excel-friendly values:
    #         - True/1   -> TRUE (Excel boolean)
    #         - False/0  -> FALSE (Excel boolean)
    #         - None/''/etc. -> "" (Empty string for blank Excel cell)
    #         - Other    -> str(val) (Text)
    #         """
    #         if val is True or (isinstance(val, (int, float)) and val == 1):
    #             return True # Excel TRUE
    #         elif val is False or (isinstance(val, (int, float)) and val == 0):
    #             return False # Excel FALSE
    #         elif val is None or val == "":
    #              return "" # Explicitly empty cell for NULL/empty
    #         else:
    #             # Handle potential string representations of booleans from inconsistent DB
    #             if isinstance(val, str):
    #                 lower_val = val.lower()
    #                 if lower_val in ('true', '1'):
    #                     return True
    #                 elif lower_val in ('false', '0'):
    #                     return False
    #             # Default: Convert to string for text fields
    #             return str(val)

    #     # Extract and format data
    #     features = paper.get('features', {})
    #     technique = paper.get('technique', {})

    #     # --- Format the 'Last Changed' date ---
    #     changed_timestamp_str = paper.get('changed', '')
    #     formatted_changed_date = ""
    #     if changed_timestamp_str:
    #         try:
    #             # Parse the ISO format timestamp
    #             dt = datetime.fromisoformat(changed_timestamp_str.replace('Z', '+00:00'))
    #             # Format as 'YYYY-MM-DD HH:MM:SS' for Excel compatibility
    #             formatted_changed_date = dt.strftime("%Y-%m-%d %H:%M:%S")
    #         except ValueError:
    #             # If parsing fails, keep the original string or leave blank
    #             formatted_changed_date = changed_timestamp_str # Or ""

    #     row_data = [
    #         paper.get('type', ''),                    # Type (text)
    #         paper.get('title', ''),                   # Title (text)
    #         paper.get('year'),                        # Year (integer)
    #         paper.get('journal', ''),                 # Journal/Conf name (text)
    #         paper.get('page_count'),                  # Pages count (integer)
    #         # --- Classification Summary ---
    #         format_excel_value(paper.get('is_offtopic')), # Off-topic (boolean/null)
    #         paper.get('relevance'),                   # Relevance (integer)
    #         format_excel_value(paper.get('is_survey')), # Survey (boolean/null)
    #         format_excel_value(paper.get('is_through_hole')), # THT (boolean/null)
    #         format_excel_value(paper.get('is_smt')),    # SMT (boolean/null)
    #         format_excel_value(paper.get('is_x_ray')),  # X-Ray (boolean/null)
    #         # --- Features Summary (Updated Order - Corrected Boolean Features) ---
    #         format_excel_value(features.get('tracks')), # Tracks (boolean/null)
    #         format_excel_value(features.get('holes')),  # Holes / Vias (boolean/null)
    #         format_excel_value(features.get('bare_pcb_other')), # Bare PCB Other (boolean/null) - ADDED
    #         format_excel_value(features.get('solder_insufficient')), # Solder Insufficient (boolean/null)
    #         format_excel_value(features.get('solder_excess')), # Solder Excess (boolean/null)
    #         format_excel_value(features.get('solder_void')), # Solder Void (boolean/null)
    #         format_excel_value(features.get('solder_crack')), # Solder Crack (boolean/null)
    #         format_excel_value(features.get('solder_other')), # Solder Other (boolean/null) - ADDED
    #         format_excel_value(features.get('missing_component')), # Missing Comp (boolean/null)
    #         format_excel_value(features.get('wrong_component')), # Wrong Comp (boolean/null)
    #         format_excel_value(features.get('orientation')), # Orientation (boolean/null)
    #         format_excel_value(features.get('component_other')), # Comp Other (boolean/null) - ADDED
    #         format_excel_value(features.get('cosmetic')), # Cosmetic (boolean/null)
    #         # Other State (boolean based on 'other' text content) - CORRECTED COMMENT
    #         format_excel_value(features.get('other') is not None and str(features.get('other', '')).strip() != ""),
    #         features.get('other', ''),               # Other Defects Text (text) - This one shows the text
    #         # --- Techniques Summary (Updated Order) ---
    #         format_excel_value(technique.get('classic_cv_based')), # Classic CV (boolean/null)
    #         format_excel_value(technique.get('ml_traditional')), # ML (boolean/null)
    #         format_excel_value(technique.get('dl_cnn_classifier')), # CNN Classifier (boolean/null)
    #         format_excel_value(technique.get('dl_cnn_detector')), # CNN Detector (boolean/null)
    #         format_excel_value(technique.get('dl_rcnn_detector')), # R-CNN Detector (boolean/null)
    #         format_excel_value(technique.get('dl_transformer')), # Transformers (boolean/null)
    #         format_excel_value(technique.get('dl_other')), # Other DL (boolean/null)
    #         format_excel_value(technique.get('hybrid')), # Hybrid (boolean/null)
    #         format_excel_value(technique.get('available_dataset')), # Datasets (boolean/null)
    #         technique.get('model', ''),              # Model name (text)
    #         # --- Metadata ---
    #         formatted_changed_date,                 # Last Changed (formatted date string)
    #         paper.get('changed_by', ''),            # Changed By (text)
    #         format_excel_value(paper.get('verified')), # Verified (boolean/null)
    #         paper.get('estimated_score'),           # Accr. Score (integer)
    #         paper.get('verified_by', ''),           # Verified By (text)
    #         # User comments state (boolean based on 'user_trace' text content) - CORRECTED COMMENT
    #         format_excel_value(paper.get('user_trace') is not None and str(paper.get('user_trace', '')).strip() != ""),
    #         paper.get('user_trace', '')             # User comments contents (text) - This one shows the text
    #     ]

    #     # Write the row data to Excel
    #     for col_num, cell_value in enumerate(row_data, 1):
    #         ws.cell(row=row_num, column=col_num, value=cell_value)

    # # Optional: Auto-adjust column widths (basic attempt)
    # for column in ws.columns:
    #     max_length = 0
    #     column_letter = column[0].column_letter # Get the column name
    #     for cell in column:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(str(cell.value))
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2)
    #     # Cap the width to prevent extremely wide columns
    #     ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

    # # Optional: Format the data as a table (requires openpyxl >= 2.5)
    # if len(papers) > 0:
    #     # Adjust the column reference to 'AQ' (assuming 37 columns now: A through AQ)
    #     # Headers are row 1, data starts row 2, so last row is len(papers) + 1
    #     tab = Table(displayName="PCBPapersTable", ref=f"A1:AQ{len(papers) + 1}")
    #     style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
    #                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    #     tab.tableStyleInfo = style
    #     ws.add_table(tab)

    # # --- NEW: Apply Conditional Formatting for Boolean Cells ---
    # # Define fills for TRUE and FALSE
    # true_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid") # Light Green
    # false_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Light Red
    # # Updated boolean column indices based on corrected new order (1-based indexing)
    # boolean_columns = [
    #     # Classification Summary
    #     6, 8, 9, 10, 11,
    #     # Features Summary (Boolean Features)
    #     12, 13, 14, # Tracks, Holes, Bare PCB Other
    #     15, 16, 17, 18, 19, # Solder Insufficient, Excess, Void, Crack, Solder Other
    #     20, 21, 22, 23, # Missing Comp, Wrong Comp, Orientation, Comp Other
    #     24, 25, 27, # Cosmetic, Other State, User Comment State
    #     # Techniques Summary
    #     28, 29, 30, 31, 32, 33, 34, 35, 36,
    #     # Metadata
    #     39 # Verified (column 39)
    # ]

    # # Iterate through rows and specified boolean columns to apply formatting
    # for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    #     for col_idx in boolean_columns:
    #         # Adjust for 0-based indexing in the row list
    #         cell = row[col_idx - 1] # col_idx is 1-based, list index is 0-based
    #         if cell.value is True:
    #             cell.fill = true_fill
    #         elif cell.value is False:
    #             cell.fill = false_fill
    #         # If cell.value is None or "", it remains unformatted (blank cell)

    # # --- Save Workbook to BytesIO object ---
    # wb.save(output)
    # output.seek(0)
    # return output.getvalue()
    return

def generate_filename(base_name, year_from, year_to, min_page_count, hide_offtopic, extra_suffix=""):
    """Generates a filename based on filters."""
    filename_parts = [base_name]
    if year_from == year_to:
        filename_parts.append(str(year_from))
    else:
        filename_parts.append(f"{year_from}-{year_to}")
    if min_page_count > 0:
        filename_parts.append(f"min{min_page_count}pg")
    if hide_offtopic:
        filename_parts.append("noOfftopic")
    if extra_suffix:
        filename_parts.append(extra_suffix)
    return "_".join(filename_parts)

def prepare_history_log_data(paper_dict, set_num=None):
    """
    Prepares llm_log data for history row rendering.
    Marks malformed entries as invalid instead of scrapping them.
    """
    # Determine which log column to use
    if set_num and set_num in [1, 2, 3]:
        raw_log = paper_dict.get(f'set_{set_num}_llm_log', '[]')
        # has_certainty = False
    else:
        raw_log = paper_dict.get('llm_log', '[]')
        # has_certainty = True
    
    try:
        log_entries = json.loads(raw_log) if raw_log else []
    except (json.JSONDecodeError, TypeError):
        log_entries = []
    
    # Required fields for classification entries
    REQUIRED_CLASSIFICATION_FIELDS = [
        'is_offtopic', 'is_survey', 'is_through_hole', 'is_smt', 'is_x_ray',
        'relevance', 'features', 'technique'
    ]
    
    # parse output JSON and validate
    for entry in log_entries:
        try:
            output_raw = entry.get('output', '{}')
            if isinstance(output_raw, str) and output_raw:
                entry['output'] = json.loads(output_raw)
            elif isinstance(output_raw, dict):
                entry['output'] = output_raw
            else:
                entry['output'] = {}
        except (json.JSONDecodeError, TypeError, AttributeError):
            entry['output'] = {}
        
        # Mark entry as valid by default, then check for malformed data
        entry['valid'] = bool(entry.get('valid', False))
        
        # Only validate classification-type entries (not verifier entries)
        if entry.get('type') in ['classifier', 'consensus', 'averaged_llm', 'user']:
            output = entry.get('output', {})
            if not isinstance(output, dict):
                entry['valid'] = False
                entry['invalid_reason'] = 'Output is not a dictionary'
            else:
                # Check for required classification fields
                missing_fields = [f for f in REQUIRED_CLASSIFICATION_FIELDS if f not in output]
                if missing_fields:
                    entry['valid'] = False
                    entry['invalid_reason'] = f'Missing required fields: {", ".join(missing_fields)}'
                else:
                    entry['valid'] = True  # Entry has all required fields
    
    # PASS 1: Ascending order - Mark changed cells (valid entries only)
    table_entries = [e for e in log_entries
                    if e.get('type') in ['classifier', 'consensus', 'averaged_llm', 'user']
                    and e.get('valid', False)]  # ← Only use valid entries for change tracking
    
    for i in range(len(table_entries)):
        current = table_entries[i]
        older = table_entries[i - 1] if i > 0 else None
        current['changed_fields'] = set()
        if older:
            older_output = older.get('output', {}) or {}
            current_output = current.get('output', {}) or {}
            if not isinstance(older_output, dict):
                older_output = {}
            if not isinstance(current_output, dict):
                current_output = {}
            for field in ['is_offtopic', 'relevance', 'is_survey',
                         'is_through_hole', 'is_smt', 'is_x_ray']:
                if current_output.get(field) != older_output.get(field):
                    current['changed_fields'].add(field)
            current_features = current_output.get('features') or {}
            older_features = older_output.get('features') or {}
            if isinstance(current_features, dict) and isinstance(older_features, dict):
                all_keys = set(current_features.keys()) | set(older_features.keys())
                for key in all_keys:
                    if current_features.get(key) != older_features.get(key):
                        current['changed_fields'].add(f'features_{key}')
            current_technique = current_output.get('technique') or {}
            older_technique = older_output.get('technique') or {}
            if isinstance(current_technique, dict) and isinstance(older_technique, dict):
                all_keys = set(current_technique.keys()) | set(older_technique.keys())
                for key in all_keys:
                    if current_technique.get(key) != older_technique.get(key):
                        current['changed_fields'].add(f'technique_{key}')
    
    changed_fields_map = {
        entry['timestamp']: entry.get('changed_fields', set())
        for entry in table_entries
    }
    
    # PASS 2: Reverse for UI, attach verifiers
    log_entries.reverse()
    processed_entries = []
    cached_verifier = None
    for entry in log_entries:
        entry_type = entry.get('type', '')
        entry['changed_fields'] = changed_fields_map.get(entry['timestamp'], set())
        if entry_type == 'verifier':
            verifier_output = entry.get('output', {}) or {}
            if not isinstance(verifier_output, dict):
                verifier_output = {}
            cached_verifier = {
                'verified': verifier_output.get('verified'),
                'estimated_score': verifier_output.get('estimated_score'),
                'verifier_trace': entry.get('trace', ''),
                'verifier_model': entry.get('model', ''),
                'verifier_timestamp': entry['timestamp']
            }
            entry['verification_data'] = None
        elif entry_type in ['classifier', 'consensus']:
            entry['verification_data'] = cached_verifier
            cached_verifier = None
        else:
            entry['verification_data'] = None
        processed_entries.append(entry)
    
    return processed_entries



# --- Jinja2-like filters ---
def render_status(value):
    """Render status value as emoji/symbol"""
    if value == 1 or value == "true" or value is True:
        return '✔️' # Checkmark for True
    elif value == 0  or value == "false" or value is False:
        return '❌' # Cross for False
    else: # None or unknown
        return '❔' # Question mark for Unknown/Null

@app.template_filter('render_status')
def render_status_filter(value):
    return render_status(value)



def render_verified_by(value):
    """
    Render verified_by value as emoji.
    Accepts the raw database value.
    Returns HTML string with emoji and tooltip if needed.
    """
    if value == 'user':
        return f'<span title="User">👤</span>' # Human emoji
    elif value is None or value == '':
        return f'<span title="Unverified">❔</span>'
    else:
        # For any other string, value is a model name, show computer emoji with tooltip
        # Escape the model name for HTML attribute safety
        escaped_model_name = str(value).replace('"', '&quot;').replace("'", "&#39;")
        return f'<span title="{escaped_model_name}">🖥️</span>'

@app.template_filter('render_verified_by')
def render_verified_by_filter(value):
    # Use Markup to tell Jinja2 that the output is safe HTML
    return Markup(render_verified_by(value)) 



def render_changed_by(value):
    """
    Render changed_by value as emoji.
    Accepts the raw database value.
    Returns HTML string with emoji and tooltip if needed.
    """
    if value == 'user':
        return f'<span title="User">👤</span>' # Human emoji
    elif value is None or value == '':
        return f'<span title="Unknown">❔</span>' # Question mark for null/empty
    else:
        # For any other string, value is a model name, show computer emoji with tooltip
        escaped_model_name = str(value).replace('"', '&quot;').replace("'", "&#39;")
        return f'<span title="{escaped_model_name}">🖥️</span>'
    
@app.template_filter('render_changed_by')
def render_changed_by_filter(value):
    # Use Markup to tell Jinja2 that the output is safe HTML
    return Markup(render_changed_by(value))



# --- BibTeX Generation Function ---
def generate_bibtex_string(paper):
    """
    Generates a raw BibTeX entry string from a paper dictionary.
    Excludes the abstract field.
    Handles different entry types and common fields.
    """
    if not paper or not paper.get('id'):
        return "Error: Paper ID missing."

    entry_type = paper.get('type', 'misc').lower() # Default to 'misc' if type is missing
    bibtex_key = paper['id'] # Use the database ID as the BibTeX key

    # Map common database fields to BibTeX fields
    field_mapping = {
        'title': paper.get('title'),
        'author': paper.get('authors'), # Assuming authors are stored as a semicolon-separated string, might need reformatting for BibTeX
        'year': paper.get('year'),
        'journal': paper.get('journal'),
        'booktitle': paper.get('journal'), # For inproceedings, conference papers - using journal field which contains conference name
        'volume': paper.get('volume'),
        'number': paper.get('number'), # Add number if it exists in DB (mapped from issue)
        'pages': paper.get('pages'), # Already normalized, might need further cleaning for BibTeX range format
        'doi': paper.get('doi'),
        'issn': paper.get('issn'),
        'month': paper.get('month'),
        'keywords': paper.get('keywords'),
        'abstract': paper.get('abstract'),  # Note: this will be excluded later
        # Add other fields as needed
        'publisher': paper.get('publisher'),  # Might be needed for books
        'institution': paper.get('institution'),  # For tech reports
        'school': paper.get('school'),  # For theses
        'chapter': paper.get('chapter'),  # For inbook entries
        'note': paper.get('note'),  # General notes
        'howpublished': paper.get('howpublished'),  # For misc entries
    }

    # Handle author formatting (semicolon-separated to ' and ' separated, potentially)
    # This is a simple conversion, might need refinement based on your exact storage format
    if field_mapping['author']:
        # Split by semicolon and join with ' and '
        authors_list = [author.strip() for author in field_mapping['author'].split(';')]
        field_mapping['author'] = ' and '.join(authors_list)

    # Handle keywords formatting (semicolon-separated to comma-separated)
    if field_mapping['keywords']:
        # Split by semicolon and join with comma for BibTeX keywords
        keywords_list = [keyword.strip() for keyword in field_mapping['keywords'].split(';')]
        field_mapping['keywords'] = ', '.join(keywords_list)

    # Handle pages formatting for BibTeX range (e.g., "123 - 125" -> "123--125")
    if field_mapping['pages']:
        # Simple replacement, might need more robust parsing if formats vary significantly
        # Remove spaces around hyphens/dashes and replace with double dash
        import re
        pages_str = field_mapping['pages']
        # Replace spaces around various dash-like characters with double hyphen
        field_mapping['pages'] = re.sub(r'\s*[-–—]\s*', '--', pages_str)

    # Determine fields relevant to the entry type and build the entry
    # This is a basic example, you might want to expand this mapping
    type_required_fields = {
        'article': ['title', 'author', 'journal', 'year'],
        'inproceedings': ['title', 'author', 'booktitle', 'year'], # Note: booktitle, not journal
        'book': ['title', 'author', 'publisher', 'year'],
        'inbook': ['title', 'author', 'chapter', 'publisher', 'year'],
        'incollection': ['title', 'author', 'booktitle', 'publisher', 'year'],
        'techreport': ['title', 'author', 'institution', 'year'],
        'phdthesis': ['title', 'author', 'school', 'year'],
        'mastersthesis': ['title', 'author', 'school', 'year'],
        'manual': ['title'],
        'misc': ['title','author', 'year'], # Default, might add author, howpublished, note
        'conference': ['title', 'author', 'booktitle', 'year'], # Treat like inproceedings
    }

    # Update the field_mapping to use the actual conference name for booktitle when type is inproceedings or conference
    if entry_type in ['inproceedings', 'conference'] and paper.get('journal'):
        field_mapping['booktitle'] = paper.get('journal')

    relevant_fields = type_required_fields.get(entry_type, ['title', 'author', 'year']) # Default fields

    # Start building the BibTeX string
    lines = [f"@{entry_type}{{{bibtex_key},"]
    for field in relevant_fields:
        value = field_mapping.get(field)
        if value is not None and str(value).strip() != "": # Only add non-empty fields
            # Basic escaping for common BibTeX characters if needed (e.g., #, {, }, $)
            # bibtex_value = str(value).replace('{', '\\{').replace('}', '\\}') # Example
            bibtex_value = str(value)
            # Ensure proper quoting, using double quotes as standard
            lines.append(f"  {field} = {{{bibtex_value}}},") # Use braces for safety

    # Add other potentially relevant fields that aren't strictly "required"
    # Include all available fields from the database schema
    other_fields = [
        'volume', 'number', 'pages', 'doi', 'issn', 'month', 'keywords',
        'publisher', 'institution', 'school', 'chapter', 'note', 'howpublished'
    ]
    for field in other_fields:
        value = field_mapping.get(field)
        if value is not None and str(value).strip() != "" and field not in relevant_fields:
             bibtex_value = str(value)
             lines.append(f"  {field} = {{{bibtex_value}}},") # Use braces

    # Remove trailing comma from the last field
    if lines and lines[-1].endswith(','):
        lines[-1] = lines[-1][:-1]

    lines.append("}") # Close the entry
    return "\n".join(lines)

# --- Register the new filter ---
@app.template_filter('bibtex')
def bibtex_filter(paper):
    """Jinja2 filter to generate a BibTeX string for a paper dictionary."""
    return generate_bibtex_string(paper)






#Routes: 
@app.route('/', methods=['GET'])
def index():
    """Main page to display the table."""
    # Get initial filter parameters from the request (or they will default inside render_papers_table)
    hide_offtopic_param = request.args.get('hide_offtopic')
    year_from_param = request.args.get('year_from')
    year_to_param = request.args.get('year_to')
    min_page_count_param = request.args.get('min_page_count')
        
    # Get the total number of papers in the database.
    conn = get_db_connection()
    total_paper_count = conn.execute("SELECT COUNT(*) FROM papers").fetchone()[0]
    conn.close()

    papers_table_content = render_papers_table(
        hide_offtopic_param=hide_offtopic_param,
        year_from_param=year_from_param,
        year_to_param=year_to_param,
        min_page_count_param=min_page_count_param,
    )
    # Pass the rendered table content and filter values to the main index template
    # Determine values to display in the input fields (use defaults if URL params were missing/invalid)
    try:
        year_from_input_value = str(int(year_from_param)) if year_from_param is not None else str(DEFAULT_YEAR_FROM)
    except ValueError:
        year_from_input_value = str(DEFAULT_YEAR_FROM)
    try:
        year_to_input_value = str(int(year_to_param)) if year_to_param is not None else str(DEFAULT_YEAR_TO)
    except ValueError:
        year_to_input_value = str(DEFAULT_YEAR_TO)
    try:
        min_page_count_input_value = str(int(min_page_count_param)) if min_page_count_param is not None else str(DEFAULT_MIN_PAGE_COUNT)
    except ValueError:
        min_page_count_input_value = str(DEFAULT_MIN_PAGE_COUNT)
    hide_offtopic_checkbox_checked = hide_offtopic_param is None or hide_offtopic_param.lower() in ['1', 'true', 'yes', 'on']

    return render_template(
        'index.html',
        papers_table_content=papers_table_content,
        hide_offtopic=hide_offtopic_checkbox_checked,
        year_from_value=year_from_input_value,
        year_to_value=year_to_input_value,
        min_page_count_value=min_page_count_input_value,
        total_paper_count=total_paper_count
    )

#Backup/restore
@app.route('/backup', methods=['GET'])
def backup_database():
    """Creates a backup of the database and related files."""
    try:
        # Create backup filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{timestamp}.parça.zst"

        # Create temporary directory for exports
        with tempfile.TemporaryDirectory() as temp_dir:
            # Generate HTML export (full, not lite)
            papers = fetch_papers(hide_offtopic=True, year_from=0, year_to=9999, min_page_count=0)
            html_content = generate_html_export_content(papers, True, 0, 9999, 0, is_lite_export=False)
            html_path = os.path.join(temp_dir, 'export.html')
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

            # Generate XLSX export
            # xlsx_content = generate_xlsx_export_content(papers)
            # xlsx_path = os.path.join(temp_dir, 'export.xlsx')
            # with open(xlsx_path, 'wb') as f:
            #     f.write(xlsx_content)

            # Create in-memory buffer for the backup
            buffer = io.BytesIO()
            
            # Create a Zstandard compressor
            cctx = zstd.ZstdCompressor(level=1, threads=-1)  # Fastest compression level
            
            # Compress the tar directly to the buffer
            with tarfile.open(fileobj=buffer, mode='w') as tar:
                # Add database file
                tar.add(DATABASE, arcname='data/new.sqlite')
                
                # Add PDF storage directory
                if os.path.exists(globals.PDF_STORAGE_DIR):
                    tar.add(globals.PDF_STORAGE_DIR, arcname='data/pdf')
                
                # Add annotated PDF storage directory
                if os.path.exists(globals.ANNOTATED_PDF_STORAGE_DIR):
                    tar.add(globals.ANNOTATED_PDF_STORAGE_DIR, arcname='data/pdf_annotated')
                
                # Add export files
                tar.add(html_path, arcname='export.html')
                # tar.add(xlsx_path, arcname='export.xlsx')
            
            # Get the uncompressed tar data
            tar_data = buffer.getvalue()
            
            # Now compress the tar data with zstd
            compressed_data = cctx.compress(tar_data)
            
            # Create a new buffer with the compressed data
            compressed_buffer = io.BytesIO(compressed_data)
            compressed_buffer.seek(0)
            
            # Create a response with the in-memory backup
            response = send_file(
                compressed_buffer,
                as_attachment=True,
                download_name=backup_filename,
                mimetype='application/zstd'
            )
            
            # Ensure the filename is set correctly in Content-Disposition
            response.headers['Content-Disposition'] = f'attachment; filename="{backup_filename}"'
            
            return response
    except Exception as e:
        print(f"Backup error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    
@app.route('/restore', methods=['POST'])
def restore_database():
    """Restores database and related files from a backup."""
    try:
        if 'backup_file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No backup file provided'}), 400

        file = request.files['backup_file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400

        if not file.filename.endswith('.parça.zst'):
            return jsonify({'status': 'error', 'message': 'Invalid backup file format. Expected .parça.zst'}), 400

        # Create temporary directory for extraction
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save uploaded file temporarily
            temp_backup_path = os.path.join(temp_dir, 'backup.parça.zst')
            file.save(temp_backup_path)

            # Decompress and extract
            dctx = zstd.ZstdDecompressor()
            with open(temp_backup_path, 'rb') as compressed_file:
                with dctx.stream_reader(compressed_file) as decomp_stream:
                    with tarfile.open(fileobj=decomp_stream, mode='r|') as tar:
                        tar.extractall(path=temp_dir)

            # Paths in the extracted archive
            extracted_db_path = os.path.join(temp_dir, 'data', 'new.sqlite')
            extracted_pdf_dir = os.path.join(temp_dir, 'data', 'pdf')
            extracted_annotated_pdf_dir = os.path.join(temp_dir, 'data', 'pdf_annotated')

            # Verify required files exist
            if not os.path.exists(extracted_db_path):
                return jsonify({'status': 'error', 'message': 'Backup does not contain required database file'}), 500

            # Backup current data before restoring (single file name, overwrites previous)
            backup_current = "backup_before_restore.parça.zst"
            backup_current_path = os.path.join(os.getcwd(), backup_current)
            cctx = zstd.ZstdCompressor(level=1)
            with cctx.stream_writer(open(backup_current_path, 'wb')) as compressor:
                with tarfile.open(fileobj=compressor, mode='w|') as tar:
                    if os.path.exists(DATABASE):
                        tar.add(DATABASE, arcname='data/new.sqlite')
                    if os.path.exists(globals.PDF_STORAGE_DIR):
                        tar.add(globals.PDF_STORAGE_DIR, arcname='data/pdf')
                    if os.path.exists(globals.ANNOTATED_PDF_STORAGE_DIR):
                        tar.add(globals.ANNOTATED_PDF_STORAGE_DIR, arcname='data/pdf_annotated')

            # Perform restoration
            # 1. Replace database
            shutil.move(extracted_db_path, DATABASE)
            
            # 2. Replace PDF directories - only if they exist in the backup
            if os.path.exists(extracted_pdf_dir):
                # Create parent directory if it doesn't exist
                os.makedirs(os.path.dirname(globals.PDF_STORAGE_DIR), exist_ok=True)
                # Remove existing directory if it exists
                if os.path.exists(globals.PDF_STORAGE_DIR):
                    shutil.rmtree(globals.PDF_STORAGE_DIR)
                # Move the extracted directory
                shutil.move(extracted_pdf_dir, globals.PDF_STORAGE_DIR)
            else:
                # Create empty PDF directory if not in backup
                os.makedirs(globals.PDF_STORAGE_DIR, exist_ok=True)
                
            if os.path.exists(extracted_annotated_pdf_dir):
                # Create parent directory if it doesn't exist
                os.makedirs(os.path.dirname(globals.ANNOTATED_PDF_STORAGE_DIR), exist_ok=True)
                # Remove existing directory if it exists
                if os.path.exists(globals.ANNOTATED_PDF_STORAGE_DIR):
                    shutil.rmtree(globals.ANNOTATED_PDF_STORAGE_DIR)
                # Move the extracted directory
                shutil.move(extracted_annotated_pdf_dir, globals.ANNOTATED_PDF_STORAGE_DIR)
            else:
                # Create empty annotated PDF directory if not in backup
                os.makedirs(globals.ANNOTATED_PDF_STORAGE_DIR, exist_ok=True)

        return jsonify({
            'status': 'success',
            'message': f'Restored successfully from backup. Previous data backed up as {backup_current}'
        })
    except Exception as e:
        print(f"Restore error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# PDF storage/annotation routes
@app.route('/upload_pdf/<paper_id>', methods=['POST']) # Removed int: converter
def upload_pdf(paper_id):
    """Handles PDF file upload for a specific paper."""
    print(f"Received upload request for paper ID: {paper_id}") # Debug print
    if 'pdf_file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part in request'}), 400

    file = request.files['pdf_file']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'No file selected'}), 400

    if file and file.filename.lower().endswith('.pdf'):
        # original_filename = secure_filename(file.filename)
        unique_filename = f"{paper_id}.pdf"
        filepath = os.path.join(globals.PDF_STORAGE_DIR, unique_filename)

        try:
            file.save(filepath)
            # Direct DB update to avoid verification reset & log pollution from update_paper_custom_fields
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE papers SET pdf_filename = ?, pdf_state = ? WHERE id = ?", (unique_filename, 'PDF', paper_id))
            conn.commit()
            conn.close()
            
            result = {'status': 'success'}
            if result['status'] == 'success':
                # Fetch updated paper data including new PDF info
                # Pass the string ID here too
                updated_paper_data = fetch_updated_paper_data(paper_id)
                if updated_paper_data['status'] == 'success':
                    # Add the PDF specific data to the response
                    updated_paper_data['pdf_filename'] = unique_filename
                    updated_paper_data['pdf_state'] = 'PDF'
                    return jsonify(updated_paper_data)
                else:
                     print(f"Failed to fetch updated data for {paper_id} after upload.")
                     return jsonify({'status': 'error', 'message': 'Failed to fetch updated paper data after upload'}), 500
            else:
                print(f"DB update failed for {paper_id} after saving file.")
                # Rollback: delete the file if DB update failed
                if os.path.exists(filepath):
                    os.remove(filepath)
                return jsonify({'status': 'error', 'message': 'Failed to update database after saving file'}), 500

        except Exception as e:
            print(f"Error saving uploaded PDF for paper {paper_id}: {e}")
            # Attempt to remove the file if saving failed partway
            if os.path.exists(filepath):
                try:
                    os.remove(filepath)
                except OSError:
                    pass # Ignore error if file removal also fails
            return jsonify({'status': 'error', 'message': 'Failed to save file'}), 500
    else:
        print(f"File type not allowed for {paper_id}: {file.filename}")
        return jsonify({'status': 'error', 'message': 'File type not allowed, only PDFs are accepted'}), 400

@app.route('/serve_pdf/<paper_id>')
def serve_pdf(paper_id):
    """
    Serves the correct PDF file (annotated or original) for the PDF.js viewer/annotator
    based on the paper_id. Also updates the pdf_state in the database
    based on the actual existence of the annotated files.
    """
    conn = get_db_connection()
    paper = conn.execute("SELECT pdf_filename, pdf_state FROM papers WHERE id = ?", (paper_id,)).fetchone()
    
    if not paper or not paper['pdf_filename']:
        conn.close()
        print(f"No PDF filename found in DB for paper_id: {paper_id}")
        abort(404)

    filename = paper['pdf_filename']
    current_db_state = paper['pdf_state']
    
    # Check for annotated and original files
    annotated_path = os.path.join(globals.ANNOTATED_PDF_STORAGE_DIR, filename)
    original_path = os.path.join(globals.PDF_STORAGE_DIR, filename)
    new_state = None
    file_to_serve = None

    if os.path.exists(annotated_path):
        # print(f"Found annotated file: {filename}")
        new_state = 'annotated'
        file_to_serve = annotated_path
    elif os.path.exists(original_path):
        # print(f"Found original file: {filename}")
        new_state = 'PDF' # Reset to 'PDF' if annotated file is missing but original exists
        file_to_serve = original_path
    else:
        print(f"No PDF file found for paper_id: {paper_id} (filename: {filename})")
        new_state = 'none'
        update_query = "UPDATE papers SET pdf_state = ? WHERE id = ?"
        conn.execute(update_query, (new_state, paper_id))
        conn.commit()
        conn.close()
        abort(404)

    if current_db_state != new_state:
        print(f"Updating pdf_state for {paper_id} from '{current_db_state}' to '{new_state}'")
        update_query = "UPDATE papers SET pdf_state = ? WHERE id = ?"
        conn.execute(update_query, (new_state, paper_id))
        conn.commit()
    else:
        print(f"pdf_state for {paper_id} is already correct ('{new_state}')")

    conn.close()

    # Use os.path.basename to get just the filename from the full path for send_from_directory
    return send_from_directory(os.path.dirname(file_to_serve), os.path.basename(file_to_serve), as_attachment=False)

@app.route('/upload_annotated_pdf/<paper_id>', methods=['POST'])
def upload_annotated_pdf(paper_id):
    """
    API call for annotator autosaving feature:
    Receives an annotated PDF file associated with a paper_id,
    saves it to the annotated storage directory, and updates the pdf_state.
    """
    if 'pdf_file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part in request'}), 400

    file = request.files['pdf_file']
    if file.filename == '' or not file.filename.lower().endswith('.pdf'):
        return jsonify({'status': 'error', 'message': 'Invalid or missing file'}), 400

    # Look up filename from paper_id to save it correctly.
    conn = get_db_connection()
    paper = conn.execute("SELECT pdf_filename FROM papers WHERE id = ?", (paper_id,)).fetchone()
    conn.close()

    if not paper or not paper['pdf_filename']:
        return jsonify({'status': 'error', 'message': f'Paper ID {paper_id} not found in DB'}), 404
    
    # Use the filename from the DB. Sanitize for security.
    filename = secure_filename(paper['pdf_filename'])
    filepath = os.path.join(globals.ANNOTATED_PDF_STORAGE_DIR, filename) # [2]

    try:
        file.save(filepath)
        # Direct DB update
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE papers SET pdf_state = ? WHERE id = ?", ('annotated', paper_id))
        conn.commit()
        conn.close()
        result = {'status': 'success'}
        if result.get('status') != 'success':
             # If DB update fails, you might want to remove the saved file to avoid inconsistency.
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'status': 'error', 'message': 'Failed to update paper state in DB'}), 500

        # print(f"Saved annotated PDF for paper {paper_id} as {filename}")
        return jsonify({'status': 'success', 'message': f'File {filename} updated successfully.'})
    except Exception as e:
        print(f"Error saving annotated PDF for paper {paper_id}: {e}")
        return jsonify({'status': 'error', 'message': 'Failed to save file on server.'}), 500

# Export routes
@app.route('/static_export', methods=['GET'])
def static_export():
    """Generate and serve a downloadable HTML snapshot based on current filters."""
    # --- Get filter parameters from the request (URL query params) ---
    hide_offtopic_param = request.args.get('hide_offtopic')
    year_from_param = request.args.get('year_from')
    year_to_param = request.args.get('year_to')
    min_page_count_param = request.args.get('min_page_count')
    # hidden params (usable, but not implemented in the Web client GUI):
    lite_param = request.args.get('lite', default='0')
    download_param = request.args.get('download', default='1')

    # --- Determine filter values ---
    hide_offtopic, year_from_value, year_to_value, min_page_count_value= get_default_filter_values(
        hide_offtopic_param, year_from_param, year_to_param, min_page_count_param
    )

    # --- Fetch papers based on these filters ---
    papers = fetch_papers(
        hide_offtopic=hide_offtopic,
        year_from=year_from_value,
        year_to=year_to_value,
        min_page_count=min_page_count_value,
    )

    is_lite_export = lite_param.lower() in ['1', 'true', 'yes']

    # --- Generate the content using the core function ---
    full_html_content = generate_html_export_content(
        papers, hide_offtopic, year_from_value, year_to_value, min_page_count_value, is_lite_export
    )

    # --- Create a filename based on filters ---
    extra_suffix = "lite" if is_lite_export else ""
    filename = generate_filename("ResearchParça", year_from_value, year_to_value, min_page_count_value, hide_offtopic, extra_suffix) + ".html"

    # --- Prepare Response Headers ---
    response_headers = {"Content-Type": "text/html"}
    if download_param == '0':       # If download=0, set Content-Disposition to 'inline' to display in browser
        response_headers["Content-Disposition"] = f"inline; filename={filename}"
        print(f"Serving static export inline: {filename}") # Optional: Log action
    else:                           # Default behavior: prompt for download
        response_headers["Content-Disposition"] = f"attachment; filename={filename}"
        print(f"Sending static export as attachment: {filename}") # Optional: Log action

    return Response(
        full_html_content,
        mimetype="text/html",
        headers=response_headers
    )

@app.route('/xlsx_export', methods=['GET'])
def export_excel():
    """Generate and serve a downloadable Excel (.xlsx) file based on current filters."""
    # --- Get filter parameters from the request (URL query params) ---
    hide_offtopic_param = request.args.get('hide_offtopic')
    year_from_param = request.args.get('year_from')
    year_to_param = request.args.get('year_to')
    min_page_count_param = request.args.get('min_page_count')

    # --- Determine filter values ---
    hide_offtopic, year_from_value, year_to_value, min_page_count_value = get_default_filter_values(
        hide_offtopic_param, year_from_param, year_to_param, min_page_count_param
    )

    # --- Fetch papers based on these filters ---
    papers = fetch_papers(
        hide_offtopic=hide_offtopic,
        year_from=year_from_value,
        year_to=year_to_value,
        min_page_count=min_page_count_value,
    )

    # --- Generate the content using the core function ---
    excel_bytes = generate_xlsx_export_content(papers)

    # --- Create a filename based on filters ---
    filename = generate_filename("ResearchParça", year_from_value, year_to_value, min_page_count_value, hide_offtopic) + ".xlsx"

    # --- Return as a downloadable attachment ---
    return Response(
        excel_bytes, # Get the bytes from the core function
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Table generation routes
@app.route('/get_detail_row', methods=['GET'])
def get_detail_row():
    """Endpoint to fetch and render the detail row content for a specific paper."""
    paper_id = request.args.get('paper_id')
    if not paper_id:
        return jsonify({'status': 'error', 'message': 'Paper ID is required'}), 400

    try:
        conn = get_db_connection()
        paper = conn.execute("SELECT * FROM papers WHERE id = ?", (paper_id,)).fetchone()
        conn.close()

        if paper:
            # Process the paper data like in fetch_papers for consistency
            paper_dict = dict(paper)
            try:
                paper_dict['features'] = json.loads(paper_dict['features'])
            except (json.JSONDecodeError, TypeError):
                paper_dict['features'] = {}
            try:
                paper_dict['technique'] = json.loads(paper_dict['technique'])
            except (json.JSONDecodeError, TypeError):
                paper_dict['technique'] = {}
            
            # Render the detail row template fragment for this specific paper
            detail_html = render_template('detail_row.html', paper=paper_dict)
            return jsonify({'status': 'success', 'html': detail_html})
        else:
            return jsonify({'status': 'error', 'message': 'Paper not found'}), 404
    except Exception as e:
        print(f"Error fetching detail row for paper {paper_id}: {e}")
        return jsonify({'status': 'error', 'message': 'Failed to fetch detail row'}), 500

app.jinja_env.filters['format_changed_timestamp'] = format_changed_timestamp


@app.route('/get_history_row', methods=['GET'])
def get_history_row():
    """Endpoint to fetch and render the history row content for a specific paper."""
    paper_id = request.args.get('paper_id')
    if not paper_id:
        return jsonify({'status': 'error', 'message': 'Paper ID is required'}), 400
    
    try:
        conn = get_db_connection()
        paper = conn.execute("SELECT * FROM papers WHERE id = ?", (paper_id,)).fetchone()
        conn.close()
        
        if paper:
            paper_dict = dict(paper)
            
            # Parse main JSON fields
            try:
                paper_dict['features'] = json.loads(paper_dict['features']) if paper_dict['features'] else {}
            except (json.JSONDecodeError, TypeError):
                paper_dict['features'] = {}
            
            try:
                paper_dict['technique'] = json.loads(paper_dict['technique']) if paper_dict['technique'] else {}
            except (json.JSONDecodeError, TypeError):
                paper_dict['technique'] = {}
            
            # Parse main_certainty
            try:
                paper_dict['main_certainty'] = json.loads(paper_dict['main_certainty']) if paper_dict['main_certainty'] else {}
            except (json.JSONDecodeError, TypeError):
                paper_dict['main_certainty'] = {}
            
            # Prepare ALL 4 logs using the SAME function
            paper_dict['llm_log_entries'] = prepare_history_log_data(paper_dict, set_num=None)
            paper_dict['set_1_llm_log_entries'] = prepare_history_log_data(paper_dict, set_num=1)
            paper_dict['set_2_llm_log_entries'] = prepare_history_log_data(paper_dict, set_num=2)
            paper_dict['set_3_llm_log_entries'] = prepare_history_log_data(paper_dict, set_num=3)
            
            # Render the history row template
            history_html = render_template('history_row.html', paper=paper_dict)
            return jsonify({'status': 'success', 'html': history_html})
        else:
            return jsonify({'status': 'error', 'message': 'Paper not found'}), 404
    except Exception as e:
        print(f"Error fetching history row for paper {paper_id}: {e}")
        return jsonify({'status': 'error', 'message': 'Failed to fetch history row'}), 50
    

@app.route('/load_table', methods=['GET'])
def load_table():
    """Endpoint to fetch and render the table content based on current filters."""
    # Get filter parameters from the request
    hide_offtopic_param = request.args.get('hide_offtopic')
    year_from_param = request.args.get('year_from')
    year_to_param = request.args.get('year_to')
    min_page_count_param = request.args.get('min_page_count')

    # Use the updated helper function to render the table, passing the search query
    table_html = render_papers_table(
        hide_offtopic_param=hide_offtopic_param,
        year_from_param=year_from_param,
        year_to_param=year_to_param,
        min_page_count_param=min_page_count_param,
    )
    return table_html

# Data import/update routes (data writing):
@app.route('/update_paper', methods=['POST'])
def update_paper():
    """Endpoint to handle AJAX updates (partial or full)."""
    data = request.get_json()
    paper_id = data.get('id')
    if not paper_id:
        return jsonify({'status': 'error', 'message': 'Paper ID is required'}), 400

    try:
        # Use 'user' as the identifier for changes made via this interface
        result = update_paper_custom_fields(paper_id, data, changed_by="user")
        # The result dict already contains status and other data
        return jsonify(result)
    except Exception as e:
        print(f"Error updating paper {paper_id}: {e}") # Log error
        return jsonify({'status': 'error', 'message': 'Failed to update database'}), 500


@app.route('/classify', methods=['POST'])
def classify_paper():
    """Endpoint to handle classification requests."""
    data = request.get_json()
    mode = data.get('mode', 'id')
    paper_id = data.get('paper_id')
    
    # Determine which queue_manager endpoint to use
    if mode == 'consensus':
        endpoint = '/consensus'
    else:
        endpoint = '/classify'
    
    # Try to call queue manager
    try:
        response = requests.post(
            f"{globals.QUEUE_MANAGER_URL}{endpoint}",
            json={'mode': mode, 'paper_id': paper_id},
            timeout=None  # No timeout for single-paper requests
        )
        if response.status_code == 200:
            result = response.json()
            if mode == 'id':
                # Single paper - fetch updated data and return
                updated_data = fetch_updated_paper_data(paper_id)
                return jsonify(updated_data)
            else:
                # Batch - return immediately
                return jsonify({'status': 'started', 'message': f'Batch classification ({mode}) initiated.'})
        else:
            return jsonify({'status': 'error', 'message': 'Queue manager returned error'}), 500
    except requests.exceptions.RequestException as e:
        # Queue manager unavailable
        return jsonify({'status': 'error', 'message': 'Queue manager unavailable'}), 503
    
@app.route('/verify', methods=['POST'])
def verify_paper():
    """Endpoint to handle verification requests."""
    data = request.get_json()
    mode = data.get('mode', 'id')
    paper_id = data.get('paper_id')
    
    try:
        response = requests.post(
            f"{globals.QUEUE_MANAGER_URL}/verify",
            json={'mode': mode, 'paper_id': paper_id},
            timeout=None
        )
        
        if response.status_code == 200:
            result = response.json()
            if mode == 'id':
                updated_data = fetch_updated_paper_data(paper_id)
                return jsonify(updated_data)
            else:
                return jsonify({'status': 'started', 'message': f'Batch verification ({mode}) initiated.'})
        else:
            return jsonify({'status': 'error', 'message': 'Queue manager returned error'}), 500
    
    except requests.exceptions.RequestException as e:
        return jsonify({'status': 'error', 'message': 'Queue manager unavailable'}), 503
    
## Outdated, must be updated to be useful for v1.2:
@app.route('/upload_bibtex', methods=['POST'])
def upload_bibtex():
    """Endpoint to handle BibTeX/CSV file upload and import."""
    global DATABASE # Assuming DATABASE is defined globally as before

    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'No selected file'}), 400

    filename = file.filename.lower()
    
    try:
        import import_bibtex 
        with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
            file.save(tmp_file.name)
            tmp_file_path = tmp_file.name

        if filename.endswith('.bib'):
            import_bibtex.import_bibtex(tmp_file_path, DATABASE)
            
        elif filename.endswith('.csv'):
            bibtex_entries = import_bibtex.convert_csv_to_bibtex(tmp_file_path)
            
            # Create temporary BibTeX file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.bib') as tmp_bib_file:
                for entry in bibtex_entries:
                    tmp_bib_file.write(entry.encode('utf-8'))
                tmp_bib_path = tmp_bib_file.name
            
            import_bibtex.import_bibtex(tmp_bib_path, DATABASE)
            
            # Clean up the temporary BibTeX file
            os.unlink(tmp_bib_path)
            
        else:
            # Clean up the temporary file before returning error
            os.unlink(tmp_file_path)
            return jsonify({'status': 'error', 'message': 'Invalid file type. Please upload a .bib or .csv file.'}), 400

        # Clean up the temporary file
        os.unlink(tmp_file_path)

        return jsonify({'status': 'success', 'message': f'{"BibTeX" if filename.endswith(".bib") else "CSV"} file imported successfully.'})

    except Exception as e:
        # Ensure cleanup even if import fails
        if 'tmp_file_path' in locals():
            try:
                os.unlink(tmp_file_path)
            except OSError:
                pass # Ignore errors during cleanup
        # Also clean up temporary BibTeX file if it was created
        if 'tmp_bib_path' in locals():
            try:
                os.unlink(tmp_bib_path)
            except OSError:
                pass
        print(f"Error importing file: {e}")
        return jsonify({'status': 'error', 'message': f'Import failed: {str(e)}'}), 500

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Browse and edit PCB inspection papers database.')
    parser.add_argument('db_file', nargs='?', help='SQLite database file path (optional)')
    args = parser.parse_args()
    
    if args.db_file:
        DATABASE = args.db_file
        print(f"Attempting to use database file from command line argument: {DATABASE}")
    elif hasattr(globals, 'DATABASE_FILE') and globals.DATABASE_FILE:
        DATABASE = globals.DATABASE_FILE
        print(f"Attempting to use database file from globals.DATABASE_FILE: {DATABASE}")

    # If DATABASE is still None (neither arg nor globals provided), or if the specified file doesn't exist,
    # fall back to 'fallback.sqlite' by copying it to globals.DATABASE_FILE location
    fallback_needed = False
    if DATABASE is None:
        fallback_needed = True
    elif not os.path.exists(DATABASE):
        fallback_needed = True
        print(f"Warning: Specified database file not found: {DATABASE}")

    if fallback_needed:
        # Check if fallback.sqlite exists in the script's directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        fallback_path = os.path.join(script_dir, 'fallback.sqlite')
        
        if not os.path.exists(fallback_path):
            print(f"Error: Fallback database file not found: {fallback_path} \nPlease ensure 'fallback.sqlite' exists in the script's directory.")
            sys.exit(1)
        
        target_database = globals.DATABASE_FILE
        
        # Copy the fallback database to the target location
        import shutil
        shutil.copy2(fallback_path, target_database)
        
        DATABASE = target_database
        print(f"Using database file: {DATABASE}")

    # Check if the final determined database file exists
    if not os.path.exists(DATABASE):
        print(f"Error: Final database file not found: {DATABASE} \nPlease provide a valid database file via command line argument, set globals.DATABASE_FILE correctly, or ensure 'fallback.sqlite' exists in the script's directory.")
        sys.exit(1) # Exit if even the fallback doesn't exist

    print(f"Starting server, database: {DATABASE}")

    # --- Open browser only once ---
    # The standard Flask/Werkzeug reloader runs the script twice:
    # 1. Once in the parent process (to manage the reloader)
    # 2. Once in the child process (the actual server, where WERKZEUG_RUN_MAIN is set)
    # We only want to open the browser in the child process.
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        # Function to open the browser after a delay
        def open_browser():
            import time
            time.sleep(2)  # Wait for the server to start
            webbrowser.open("http://127.0.0.1:5000")

        # Start the browser opener in a separate thread
        threading.Thread(target=open_browser, daemon=True).start()
        print(" * Visit http://127.0.0.1:5000 to view the table.")
    
    app.run(host='0.0.0.0', port=5000, debug=True)